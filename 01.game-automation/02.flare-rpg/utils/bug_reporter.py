# 역할: 테스트 실패 정보를 CSV/XLSX 버그 리포트로 저장
# 02.tools 의 컬럼 구조와 호환

import os
import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook

# 결과 저장 폴더
RESULTS_DIR = "bug_reports"

# Allure severity → 심각도/우선순위 매핑
SEVERITY_MAP = {
    "blocker":  ("Critical", "P1"),
    "critical": ("High",     "P2"),
    "normal":   ("Medium",   "P3"),
    "minor":    ("Low",      "P4"),
    "trivial":  ("Low",      "P4"),
}

# 버그 리포트 컬럼
COLUMNS = [
    "버그ID", "TC_ID", "제목", "심각도", "우선순위", "플랫폼",
    "버전", "상태", "발견자", "발견일", "해결일", "재현율"
]


class BugReporter:
    """
    테스트 실패 시 버그 리포트 CSV/XLSX 생성
    같은 실행 세션 내 여러 실패를 누적 기록
    """

    def __init__(self):
        os.makedirs(RESULTS_DIR, exist_ok=True)
        # 세션 시작 시각 (파일명에 사용)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.csv_path = os.path.join(RESULTS_DIR, f"bugs_{self.timestamp}.csv")
        self.xlsx_path = os.path.join(RESULTS_DIR, f"bugs_{self.timestamp}.xlsx")
        # 버그 ID 자동 증가용 카운터
        self.bug_count = 0

    def add_bug(self, test_id, test_name, severity="critical"):
        """
        버그 정보 추가
        CSV 와 XLSX 둘 다 한 번에 기록

        매개변수:
        - test_id: 테스트 ID (예: "FR-008")
        - test_name: 테스트 이름
        - severity: Allure severity (소문자: blocker/critical/normal/minor)
        """
        self.bug_count += 1
        bug_id = f"BUG-{self.bug_count:03d}"

        # severity 매핑 (소문자 처리)
        sev_key = severity.lower() if severity else "critical"
        severity_kr, priority = SEVERITY_MAP.get(sev_key, ("High", "P2"))

        # 버그 정보
        bug_row = {
            "버그ID":   bug_id,
            "TC_ID":   test_id,
            "제목":     test_name,
            "심각도":   severity_kr,
            "우선순위": priority,
            "플랫폼":   "PC",
            "버전":     "Flare RPG v1.15",
            "상태":     "미해결",
            "발견자":   "QA Bot",
            "발견일":   datetime.now().strftime("%Y-%m-%d"),
            "해결일":   "",
            "재현율":   "100%",
        }

        # CSV 기록 (utf-8-sig: 엑셀에서 한글 안 깨짐)
        # 첫 버그면 헤더부터 작성, 이후엔 append
        write_header = not os.path.exists(self.csv_path)
        with open(self.csv_path, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS)
            if write_header:
                writer.writeheader()
            writer.writerow(bug_row)

        # XLSX 기록
        if not os.path.exists(self.xlsx_path):
            # 첫 버그면 새 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "버그 리포트"
            ws.append(COLUMNS)
        else:
            # 이후엔 기존 파일 열어서 추가
            wb = load_workbook(self.xlsx_path)
            ws = wb.active

        ws.append([bug_row[col] for col in COLUMNS])
        wb.save(self.xlsx_path)

        return bug_id

    def get_paths(self):
        """생성된 리포트 파일 경로 반환"""
        return {
            "csv":  self.csv_path if os.path.exists(self.csv_path) else None,
            "xlsx": self.xlsx_path if os.path.exists(self.xlsx_path) else None,
        }