# 역할: 테스트 실패 정보를 CSV/XLSX 버그 리포트로 저장

import csv
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook


# ── 1. 변수 선언부 ──────────────────────────────────

RESULTS_DIR = "bug_reports"

# Allure severity → (심각도, 우선순위) 매핑
SEVERITY_MAP = {
    "blocker": ("Critical", "P1"),
    "critical": ("High", "P2"),
    "normal": ("Medium", "P3"),
    "minor": ("Low", "P4"),
    "trivial": ("Low", "P4"),
}

DEFAULT_SEVERITY = ("High", "P2")

# 02.tools 의 버그 리포트 컬럼 구조와 호환
COLUMNS = [
    "버그ID", "TC_ID", "제목", "심각도", "우선순위", "플랫폼",
    "버전", "상태", "발견자", "발견일", "해결일", "재현율",
]


# ── 2. 함수 선언부 ──────────────────────────────────

class BugReporter:
    """실패한 테스트를 CSV/XLSX 버그 리포트로 누적 기록."""

    def __init__(self, device=None):
        """device: uiautomator2 객체. 플랫폼 컬럼에 기기 정보를 넣는 데 사용."""
        os.makedirs(RESULTS_DIR, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.csv_path = os.path.join(RESULTS_DIR, f"bugs_{timestamp}.csv")
        self.xlsx_path = os.path.join(RESULTS_DIR, f"bugs_{timestamp}.xlsx")
        self.bug_count = 0
        self.platform = self._platform_label(device)

    def _platform_label(self, device):
        """플랫폼 컬럼에 들어갈 문자열 생성."""
        if device is None:
            return "Android"

        info = device.device_info
        return f"Android {info['version']} / {info['model']}"

    def _build_row(self, test_id, test_name, severity):
        """버그 리포트 한 행 조립."""
        self.bug_count += 1
        severity_kr, priority = SEVERITY_MAP.get(severity.lower(), DEFAULT_SEVERITY)

        return {
            "버그ID": f"BUG-{self.bug_count:03d}",
            "TC_ID": test_id,
            "제목": test_name,
            "심각도": severity_kr,
            "우선순위": priority,
            "플랫폼": self.platform,
            "버전": "Privacy Friendly 2048",
            "상태": "미해결",
            "발견자": "QA Bot",
            "발견일": datetime.now().strftime("%Y-%m-%d"),
            "해결일": "",
            "재현율": "100%",
        }

    def _write_csv(self, row):
        """CSV 에 한 행 추가. 첫 호출 시 헤더도 작성."""
        write_header = not os.path.exists(self.csv_path)

        # utf-8-sig 로 저장해야 엑셀에서 한글이 깨지지 않음
        with open(self.csv_path, "a", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=COLUMNS)
            if write_header:
                writer.writeheader()
            writer.writerow(row)

    def _write_xlsx(self, row):
        """XLSX 에 한 행 추가. 첫 호출 시 워크북 생성."""
        if os.path.exists(self.xlsx_path):
            workbook = load_workbook(self.xlsx_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "버그 리포트"
            sheet.append(COLUMNS)

        sheet.append([row[column] for column in COLUMNS])
        workbook.save(self.xlsx_path)

    def add_bug(self, test_id, test_name, severity="critical"):
        """버그 한 건을 CSV/XLSX 양쪽에 기록하고 버그 ID 반환."""
        row = self._build_row(test_id, test_name, severity)
        self._write_csv(row)
        self._write_xlsx(row)
        return row["버그ID"]

    def get_paths(self):
        """생성된 리포트 파일 경로 반환. 미생성 시 None."""
        return {
            "csv": self.csv_path if os.path.exists(self.csv_path) else None,
            "xlsx": self.xlsx_path if os.path.exists(self.xlsx_path) else None,
        }