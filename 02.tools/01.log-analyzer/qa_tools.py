# 역할: 로그 파일 분석에 필요한 핵심 함수 모음
# 다른 모듈(main.py)에서 import해서 사용

import re                  # 정규표현식 - 버그ID/시간 패턴 추출
import json                # JSON 파일 입출력 - 히스토리 저장/불러오기
from pathlib import Path   # 경로 관리 - 파일/폴더 탐색 및 생성
from openpyxl import Workbook                          # 엑셀 파일 생성
from openpyxl.styles import PatternFill, Font, Alignment  # 엑셀 스타일
from openpyxl.utils import get_column_letter           # 열 이름 변환
from datetime import datetime                          # 날짜/시간
from collections import Counter                        # 중복 카운팅

# qa_tools.py 가 있는 폴더를 기준 경로로 설정
# → 어디서 실행해도 파일을 올바른 위치에서 찾음
기준경로 = Path(__file__).parent


# ────────────────────────────────────────
# ① 로그 파일 읽기
# ────────────────────────────────────────
def 로그_읽기(파일명):
    """
    로그 파일을 읽어서 줄 목록으로 반환

    매개변수:
    - 파일명: 읽을 파일 이름 (예: "bug_log.txt")
              기준경로(qa_tools.py 위치) 기준으로 찾음

    반환값:
    - 줄 목록 (list): 파일의 각 줄을 담은 리스트
    - [] (빈 리스트): 파일이 없을 때

    지원 형식:
    - UTF-8 인코딩 텍스트 파일 (.txt)
    - 한 줄에 하나의 로그 항목
    """
    경로 = 기준경로 / 파일명
    if not 경로.exists():
        print("❌ 파일이 없어요:", 경로)
        return []
    with open(경로, "r", encoding="utf-8") as f:
        return f.readlines()


# ────────────────────────────────────────
# ② 에러 로그만 필터링
# ────────────────────────────────────────
def 에러_필터링(로그목록):
    """
    전체 로그에서 ERROR 가 포함된 줄만 추출

    매개변수:
    - 로그목록: 로그_읽기()로 읽은 줄 목록

    반환값:
    - 에러목록 (list): ERROR 포함 줄만 담은 리스트

    동작:
    - 줄 앞뒤 공백/줄바꿈 제거 (strip)
    - "ERROR" 문자열 포함 여부로 필터링
    - 대소문자 구분 있음 ("error" 는 필터링 안 됨)
    """
    에러목록 = []
    for 로그 in 로그목록:
        로그 = 로그.strip()          # 줄 앞뒤 공백/줄바꿈 제거
        if "ERROR" in 로그:          # ERROR 포함 여부 확인
            에러목록.append(로그)
    return 에러목록


# ────────────────────────────────────────
# ③ 버그 ID / 시간 추출
# ────────────────────────────────────────
def 버그정보_추출(로그):
    """
    로그 한 줄에서 버그 ID와 발생 시간을 추출

    매개변수:
    - 로그: 분석할 로그 문자열
            예) "ERROR: 캐릭터 충돌 BUG-001 at 14:30:22"

    반환값:
    - 딕셔너리: {"로그": 원본, "버그ID": "BUG-001", "시간": "14:30:22"}
    - 패턴 없으면 "없음" 으로 채워서 반환

    정규표현식 패턴:
    - 버그ID: BUG- 뒤에 오는 숫자 1개 이상 (BUG-001, BUG-42 등)
    - 시간  : 00:00:00 형식 (시:분:초 각 2자리)
    """
    버그ID = re.search(r"BUG-\d+", 로그)            # BUG-숫자 패턴 찾기
    시간   = re.search(r"\d{2}:\d{2}:\d{2}", 로그)  # HH:MM:SS 패턴 찾기
    return {
        "로그"  : 로그,
        "버그ID": 버그ID.group() if 버그ID else "없음",  # 없으면 "없음"
        "시간"  : 시간.group()   if 시간   else "없음"
    }


# ────────────────────────────────────────
# ④ 심각도 자동 분류
# ────────────────────────────────────────
def 심각도_분류(로그):
    """
    로그 내용의 키워드로 심각도를 자동 판단

    매개변수:
    - 로그: 분석할 로그 문자열

    반환값:
    - "Critical" / "Major" / "Minor" 중 하나
    - 키워드 없으면 기본값 "Minor" 반환

    키워드 매핑:
    - Critical: 충돌, crash, 서버 다운, 응답 없음
    - Major   : 프레임 드랍, 렉, 지연, 오류
    - Minor   : 경고, warning, 메모리
    """
    키워드맵 = {
        "Critical": ["충돌", "crash", "서버 다운", "응답 없음"],
        "Major"   : ["프레임 드랍", "렉", "지연", "오류"],
        "Minor"   : ["경고", "warning", "메모리"],
    }
    for 심각도, 키워드목록 in 키워드맵.items():
        for 키워드 in 키워드목록:
            if 키워드.lower() in 로그.lower():  # 대소문자 구분 없이 비교
                return 심각도
    return "Minor"  # 매칭 키워드 없으면 기본값


# ────────────────────────────────────────
# ⑤ 중복 버그 감지
# ────────────────────────────────────────
def 중복_감지(버그목록):
    """
    같은 버그 ID가 2회 이상 발생한 항목 탐지

    매개변수:
    - 버그목록: 버그정보_추출()로 만든 딕셔너리 리스트

    반환값:
    - 딕셔너리: {버그ID: 발생횟수} 형태
               2회 이상인 것만 포함
    - {} (빈 딕셔너리): 중복 없을 때

    동작:
    - "없음" 버그ID는 카운팅에서 제외
    - Counter로 각 ID의 발생 횟수 집계
    """
    ID목록  = [버그["버그ID"] for 버그 in 버그목록 if 버그["버그ID"] != "없음"]
    카운터  = Counter(ID목록)
    중복목록 = {id: 수 for id, 수 in 카운터.items() if 수 > 1}
    return 중복목록


# ────────────────────────────────────────
# ⑥ 이전 결과와 비교
# ────────────────────────────────────────
def 히스토리_비교(현재버그목록):
    """
    이전 실행 결과와 비교해서 신규/해결 버그 추적

    매개변수:
    - 현재버그목록: 이번 실행에서 발견된 버그 목록

    반환값:
    - 딕셔너리: {"신규": [새 버그ID 목록], "해결": [사라진 버그ID 목록]}

    동작:
    - 결과/bug_history.json 에서 이전 기록 불러오기
    - 이전 기록 없으면 현재를 저장하고 신규로 처리
    - 현재에만 있는 ID → 신규 버그
    - 이전에만 있는 ID → 해결된 버그
    - 비교 후 현재 기록을 저장해서 다음 실행 때 사용
    """
    히스토리파일 = 기준경로 / "결과" / "bug_history.json"
    현재ID목록   = set(버그["버그ID"] for 버그 in 현재버그목록)

    # 이전 기록 없으면 현재를 저장하고 신규로 처리
    if not 히스토리파일.exists():
        히스토리_저장(현재버그목록)
        return {"신규": list(현재ID목록), "해결": []}

    # 이전 기록 불러오기
    with open(히스토리파일, "r", encoding="utf-8") as f:
        이전데이터 = json.load(f)

    이전ID목록 = set(이전데이터.get("버그ID목록", []))

    신규버그 = list(현재ID목록 - 이전ID목록)  # 현재에만 있는 것
    해결버그 = list(이전ID목록 - 현재ID목록)  # 이전에만 있는 것

    히스토리_저장(현재버그목록)  # 현재 기록 저장
    return {"신규": 신규버그, "해결": 해결버그}


def 히스토리_저장(버그목록):
    """
    현재 버그 목록을 JSON 파일로 저장
    다음 실행 때 히스토리_비교()가 이 파일을 불러옴

    저장 위치: 결과/bug_history.json
    저장 형식: {"날짜": "YYYY-MM-DD HH:MM:SS", "버그ID목록": [...]}
    """
    결과폴더    = 기준경로 / "결과"
    결과폴더.mkdir(exist_ok=True)

    히스토리파일 = 결과폴더 / "bug_history.json"
    오늘        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    데이터 = {
        "날짜"     : 오늘,
        "버그ID목록": [버그["버그ID"] for 버그 in 버그목록]
    }

    with open(히스토리파일, "w", encoding="utf-8") as f:
        json.dump(데이터, f, ensure_ascii=False, indent=4)


# ────────────────────────────────────────
# ⑦ 열 너비 자동 조정
# ────────────────────────────────────────
def 열너비_조정(ws):
    """
    워크시트의 각 열 너비를 내용 길이에 맞게 자동 조정

    매개변수:
    - ws: openpyxl 워크시트 객체

    동작:
    - 각 열의 모든 셀 중 가장 긴 내용 기준으로 너비 설정
    - 여백 4 추가해서 답답하지 않게
    - MergedCell(병합 셀)은 try/except로 건너뜀
    - get_column_letter 사용으로 병합 셀 에러 방지
    """
    for i, 열 in enumerate(ws.columns, 1):
        최대길이 = 0
        열이름  = get_column_letter(i)  # 인덱스로 열 이름 가져오기 (A, B, C...)

        for 셀 in 열:
            try:
                if 셀.value:
                    최대길이 = max(최대길이, len(str(셀.value)))
            except:
                pass  # 병합 셀은 건너뜀

        ws.column_dimensions[열이름].width = 최대길이 + 4


# ────────────────────────────────────────
# ⑧ 엑셀 저장
# ────────────────────────────────────────
def 엑셀_저장(버그목록, 전체로그, 비교결과, 중복결과):
    """
    분석 결과를 5시트 엑셀 파일로 저장

    매개변수:
    - 버그목록  : 버그정보_추출()로 만든 딕셔너리 리스트
    - 전체로그  : 로그_읽기()로 읽은 전체 줄 목록
    - 비교결과  : 히스토리_비교()의 반환값
    - 중복결과  : 중복_감지()의 반환값

    반환값:
    - 저장된 파일 경로 (Path 객체)

    저장 위치: 결과/bug_report_YYYY-MM-DD.xlsx

    시트 구성:
    - 전체 로그   : ERROR🔴 / WARNING🟡 / INFO🟢 색상 구분
    - 버그 리포트 : Critical🔴 / Major🟡 / Minor🟢 심각도 분류
    - 통계 요약   : 유형별 건수
    - 변경 사항   : 신규🔴 / 해결🟢 버그
    - 중복 버그   : 반복 발생 버그🟡
    """
    결과폴더 = 기준경로 / "결과"
    결과폴더.mkdir(exist_ok=True)

    오늘   = datetime.now().strftime("%Y-%m-%d")
    파일명 = 결과폴더 / f"bug_report_{오늘}.xlsx"

    wb = Workbook()

    # ── 시트 1: 전체 로그 ──
    ws1 = wb.active
    ws1.title = "전체 로그"
    ws1.append(["로그 유형", "내용"])

    for 셀 in ws1[1]:
        셀.fill      = PatternFill("solid", fgColor="4472C4")
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")

    색상맵 = {
        "ERROR":   "FF0000",
        "WARNING": "FFC000",
        "INFO":    "70AD47",
    }

    for 로그 in 전체로그:
        로그 = 로그.strip()
        유형 = "INFO"
        if "ERROR"   in 로그: 유형 = "ERROR"
        elif "WARNING" in 로그: 유형 = "WARNING"

        ws1.append([유형, 로그])
        현재행 = ws1.max_row
        for 셀 in ws1[현재행]:
            셀.fill = PatternFill("solid", fgColor=색상맵[유형])

    # ── 시트 2: 버그 리포트 ──
    ws2 = wb.create_sheet(title="버그 리포트")
    ws2.append(["버그ID", "시간", "심각도", "로그 내용"])

    for 셀 in ws2[1]:
        셀.fill      = PatternFill("solid", fgColor="4472C4")
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")

    심각도색상 = {
        "Critical": "FF0000",
        "Major"   : "FFC000",
        "Minor"   : "70AD47",
    }

    for 버그 in 버그목록:
        심각도 = 심각도_분류(버그["로그"])
        ws2.append([버그["버그ID"], 버그["시간"], 심각도, 버그["로그"]])
        현재행 = ws2.max_row
        for 셀 in ws2[현재행]:
            셀.fill = PatternFill("solid", fgColor=심각도색상[심각도])
            셀.font = Font(color="FFFFFF")

    # ── 시트 3: 통계 요약 ──
    ws3 = wb.create_sheet(title="통계 요약")
    ws3.append(["항목", "건수"])

    for 셀 in ws3[1]:
        셀.fill      = PatternFill("solid", fgColor="4472C4")
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")

    통계데이터 = [
        ("ERROR",   sum(1 for 로그 in 전체로그 if "ERROR"   in 로그), "FF0000"),
        ("WARNING", sum(1 for 로그 in 전체로그 if "WARNING" in 로그), "FFC000"),
        ("INFO",    sum(1 for 로그 in 전체로그 if "INFO"    in 로그), "70AD47"),
        ("전체",    len(전체로그),                                     "4472C4"),
    ]

    for 항목, 건수, 색상 in 통계데이터:
        ws3.append([항목, 건수])
        현재행 = ws3.max_row
        for 셀 in ws3[현재행]:
            셀.fill      = PatternFill("solid", fgColor=색상)
            셀.font      = Font(color="FFFFFF", bold=True)
            셀.alignment = Alignment(horizontal="center")

    # ── 시트 4: 변경 사항 ──
    ws4 = wb.create_sheet(title="변경 사항")
    ws4.append(["구분", "버그ID"])

    for 셀 in ws4[1]:
        셀.fill      = PatternFill("solid", fgColor="4472C4")
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")

    for id in 비교결과["신규"]:
        ws4.append(["🆕 신규", id])
        현재행 = ws4.max_row
        for 셀 in ws4[현재행]:
            셀.fill = PatternFill("solid", fgColor="FF0000")
            셀.font = Font(color="FFFFFF")

    for id in 비교결과["해결"]:
        ws4.append(["✅ 해결", id])
        현재행 = ws4.max_row
        for 셀 in ws4[현재행]:
            셀.fill = PatternFill("solid", fgColor="70AD47")
            셀.font = Font(color="FFFFFF")

    # ── 시트 5: 중복 버그 ──
    ws5 = wb.create_sheet(title="중복 버그")
    ws5.append(["버그ID", "발생 횟수"])

    for 셀 in ws5[1]:
        셀.fill      = PatternFill("solid", fgColor="4472C4")
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")

    for id, 횟수 in 중복결과.items():
        ws5.append([id, 횟수])
        현재행 = ws5.max_row
        for 셀 in ws5[현재행]:
            셀.fill      = PatternFill("solid", fgColor="FFC000")
            셀.font      = Font(color="FFFFFF", bold=True)
            셀.alignment = Alignment(horizontal="center")

    # 모든 시트 열 너비 자동 조정
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        열너비_조정(ws)

    wb.save(파일명)
    return 파일명