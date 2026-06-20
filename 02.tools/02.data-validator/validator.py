# 역할: CSV/XLSX 파일의 데이터 무결성을 자동으로 검사하는 모듈
# 검사 규칙은 rules/ 폴더의 JSON 파일로 관리
# 공통 모듈: common.file_io, common.excel_style 사용
#
# 검사 항목:
#   - 필수 컬럼 존재 여부
#   - 필수 값 누락 (빈 값)
#   - 허용되지 않는 값 (오타)
#   - 날짜 형식 오류
#   - 중복 데이터
#   - 패턴 오류 (정규표현식)
#   - 공백 오류

import sys
import re
import json
import pandas as pd
from pathlib import Path
from datetime import datetime

# 02.tools 폴더를 Python 경로에 추가
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# 공통 모듈
from common.file_io import (
    파일_읽기, 결과폴더_생성, JSON_쓰기, HTML_쓰기,
    타임스탬프, Latest_복사
)
from common.excel_style import (
    헤더_스타일, 행_색상, 색상_가져오기, 열너비_조정, 색상
)


# validator.py 가 있는 폴더를 기준 경로로 설정
기준경로 = Path(__file__).parent


# ────────────────────────────────────────
# ① 규칙 파일 로딩 (자동 매핑)
# ────────────────────────────────────────
def 규칙_로딩(파일명):
    """
    CSV 파일 이름에 맞는 규칙 JSON 파일을 자동으로 찾아서 로딩

    규칙 파일 매핑:
    - 파일명에 "bugs" 포함     → rules/bugs.json
    - 파일명에 "testcases" 포함 → rules/testcases.json
    - 파일명에 "items" 포함     → rules/items.json (기획 데이터)
    - 파일명에 "skills" 포함    → rules/skills.json (기획 데이터)
    - 파일명에 "monsters" 포함  → rules/monsters.json (기획 데이터)
    - 매칭 안 되면             → rules/default.json
    - default.json 도 없으면   → 내장 기본 규칙

    반환값:
    - 딕셔너리: 규칙 데이터
    """
    규칙폴더 = 기준경로 / "rules"
    파일명소문자 = str(파일명).lower()

    # 파일명으로 규칙 파일 자동 매핑
    규칙파일 = None
    매핑목록 = {
        "bugs"     : "bugs.json",
        "testcases": "testcases.json",
        "items"    : "items.json",
        "skills"   : "skills.json",
        "monsters" : "monsters.json",
        "quests"   : "quests.json",
    }

    for 키워드, 파일 in 매핑목록.items():
        if 키워드 in 파일명소문자:
            규칙파일 = 규칙폴더 / 파일
            break

    # 매핑 안 되면 default.json
    if 규칙파일 is None or not 규칙파일.exists():
        규칙파일 = 규칙폴더 / "default.json"

    # default.json 도 없으면 내장 기본 규칙
    if not 규칙파일.exists():
        print("[WARN]  규칙 파일 없음 → 기본 규칙으로 검사")
        return {
            "설명": "내장 기본 규칙",
            "필수_컬럼": [],
            "컬럼_규칙": {}
        }

    with open(규칙파일, "r", encoding="utf-8") as f:
        규칙 = json.load(f)

    print(f"규칙 파일 로딩: {규칙파일.name} ({규칙.get('설명', '')})")
    return 규칙


# ────────────────────────────────────────
# ② 데이터 읽기
# ────────────────────────────────────────
def 데이터_읽기(파일명):
    """
    CSV/XLSX 파일을 읽어서 DataFrame 반환
    common.file_io.파일_읽기 사용
    """
    df = 파일_읽기(파일명, 기준경로)
    if df is None:
        return None
    return df


# ────────────────────────────────────────
# ③ 필수 컬럼 검사
# ────────────────────────────────────────
def 필수컬럼_검사(df, 규칙):
    """
    필수 컬럼이 CSV에 존재하는지 검사
    """
    오류목록 = []
    필수컬럼 = 규칙.get("필수_컬럼", [])

    for 컬럼 in 필수컬럼:
        if 컬럼 not in df.columns:
            오류목록.append({
                "유형": "필수컬럼",
                "행": "-",
                "컬럼": 컬럼,
                "값": "-",
                "내용": f"필수 컬럼 '{컬럼}' 이 없습니다"
            })

    return 오류목록


# ────────────────────────────────────────
# ④ 빈 값 검사
# ────────────────────────────────────────
def 빈값_검사(df, 규칙):
    """
    필수 항목에 빈 값이 있는지 검사
    규칙 파일의 "필수": true 인 컬럼만 검사
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    if 컬럼규칙:
        검사컬럼 = [컬럼 for 컬럼, 설정 in 컬럼규칙.items()
                    if 설정.get("필수", False) and 컬럼 in df.columns]
    else:
        검사컬럼 = list(df.columns)

    for 컬럼 in 검사컬럼:
        빈행목록 = df[df[컬럼].isna() | (df[컬럼].astype(str).str.strip() == "")]
        for idx in 빈행목록.index:
            오류목록.append({
                "유형": "빈값",
                "행": idx + 2,
                "컬럼": 컬럼,
                "값": "(비어있음)",
                "내용": f"{idx+2}행 '{컬럼}' 값이 비어있습니다"
            })

    return 오류목록


# ────────────────────────────────────────
# ⑤ 허용값 검사
# ────────────────────────────────────────
def 허용값_검사(df, 규칙):
    """
    특정 컬럼의 값이 허용된 값 목록에 포함되는지 검사
    오타 감지에 핵심 역할
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    for 컬럼, 설정 in 컬럼규칙.items():
        허용값 = 설정.get("허용값", None)
        if 허용값 is None or 컬럼 not in df.columns:
            continue

        for idx, 행 in df.iterrows():
            값 = str(행[컬럼]).strip()
            if 값 == "" or 값 == "nan":
                continue
            if 값 not in 허용값:
                오류목록.append({
                    "유형": "허용값",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": 값,
                    "내용": f"{idx+2}행 '{컬럼}' 값 '{값}' 은 허용되지 않음 (허용값: {', '.join(허용값)})"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑥ 날짜 형식 검사
# ────────────────────────────────────────
def 날짜형식_검사(df, 규칙):
    """
    날짜 컬럼의 값이 올바른 형식인지 검사
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    for 컬럼, 설정 in 컬럼규칙.items():
        날짜형식 = 설정.get("날짜형식", None)
        if 날짜형식 is None or 컬럼 not in df.columns:
            continue

        for idx, 행 in df.iterrows():
            값 = str(행[컬럼]).strip()
            if 값 == "" or 값 == "nan":
                continue
            try:
                datetime.strptime(값, 날짜형식)
            except ValueError:
                오류목록.append({
                    "유형": "날짜형식",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": 값,
                    "내용": f"{idx+2}행 '{컬럼}' 값 '{값}' 은 올바른 날짜 형식이 아님 (형식: {날짜형식})"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑦ 중복 데이터 검사
# ────────────────────────────────────────
def 중복_검사(df, 규칙):
    """
    중복불가로 지정된 컬럼에 중복 값이 있는지 검사
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    중복검사컬럼 = [컬럼 for 컬럼, 설정 in 컬럼규칙.items()
                    if 설정.get("중복불가", False) and 컬럼 in df.columns]

    if 중복검사컬럼:
        for 컬럼 in 중복검사컬럼:
            중복 = df[df[컬럼].duplicated(keep=False)]
            중복값목록 = 중복[컬럼].unique()
            for 값 in 중복값목록:
                중복행 = df[df[컬럼] == 값].index.tolist()
                행번호 = ", ".join([str(i+2) for i in 중복행])
                오류목록.append({
                    "유형": "중복",
                    "행": 행번호,
                    "컬럼": 컬럼,
                    "값": str(값),
                    "내용": f"'{컬럼}' 값 '{값}' 이 중복됨 ({행번호}행)"
                })
    else:
        중복행 = df[df.duplicated(keep=False)]
        if len(중복행) > 0:
            오류목록.append({
                "유형": "중복",
                "행": "-",
                "컬럼": "전체",
                "값": f"{len(중복행)}건",
                "내용": f"완전히 동일한 행이 {len(중복행)}건 발견됨"
            })

    return 오류목록


# ────────────────────────────────────────
# ⑧ 패턴 검사 (정규표현식)
# ────────────────────────────────────────
def 패턴_검사(df, 규칙):
    """
    정규표현식 패턴에 맞지 않는 값 검사
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    for 컬럼, 설정 in 컬럼규칙.items():
        패턴 = 설정.get("패턴", None)
        if 패턴 is None or 컬럼 not in df.columns:
            continue

        for idx, 행 in df.iterrows():
            값 = str(행[컬럼]).strip()
            if 값 == "" or 값 == "nan":
                continue
            if not re.fullmatch(패턴, 값):
                오류목록.append({
                    "유형": "패턴",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": 값,
                    "내용": f"{idx+2}행 '{컬럼}' 값 '{값}' 이 패턴에 맞지 않음 (패턴: {패턴})"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑨ 숫자 범위 검사 (기획 데이터용)
# ────────────────────────────────────────
def 숫자범위_검사(df, 규칙):
    """
    숫자 값이 지정된 범위 안에 있는지 검사
    기획 데이터(HP, 가격, 공격력 등)에 유용

    규칙 예시:
    "HP": {"최소값": 1, "최대값": 99999}
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    for 컬럼, 설정 in 컬럼규칙.items():
        최소값 = 설정.get("최소값", None)
        최대값 = 설정.get("최대값", None)
        if (최소값 is None and 최대값 is None) or 컬럼 not in df.columns:
            continue

        for idx, 행 in df.iterrows():
            값 = 행[컬럼]
            if pd.isna(값):
                continue
            try:
                숫자값 = float(값)
                if 최소값 is not None and 숫자값 < 최소값:
                    오류목록.append({
                        "유형": "범위",
                        "행": idx + 2,
                        "컬럼": 컬럼,
                        "값": str(값),
                        "내용": f"{idx+2}행 '{컬럼}' 값 {값} 이 최소값 {최소값} 미만"
                    })
                if 최대값 is not None and 숫자값 > 최대값:
                    오류목록.append({
                        "유형": "범위",
                        "행": idx + 2,
                        "컬럼": 컬럼,
                        "값": str(값),
                        "내용": f"{idx+2}행 '{컬럼}' 값 {값} 이 최대값 {최대값} 초과"
                    })
            except (ValueError, TypeError):
                오류목록.append({
                    "유형": "범위",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": str(값),
                    "내용": f"{idx+2}행 '{컬럼}' 값 '{값}' 이 숫자가 아님"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑩ 공백 오류 검사
# ────────────────────────────────────────
def 공백_검사(df):
    """
    모든 컬럼의 값에서 앞뒤 공백 오류 검사
    """
    오류목록 = []

    for 컬럼 in df.columns:
        for idx, 행 in df.iterrows():
            값 = 행[컬럼]
            if pd.isna(값):
                continue
            값문자열 = str(값)
            if 값문자열 != 값문자열.strip():
                오류목록.append({
                    "유형": "공백",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": repr(값문자열),
                    "내용": f"{idx+2}행 '{컬럼}' 값에 앞뒤 공백이 있음"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑪ 검증 리포트 저장
# ────────────────────────────────────────
def 리포트_저장(오류목록, 파일명):
    """
    검증 결과를 엑셀 파일로 저장
    공통 모듈 사용: 헤더_스타일, 행_색상, 열너비_조정
    """
    결과폴더 = 결과폴더_생성(기준경로)

    시각   = 타임스탬프()
    저장명 = 결과폴더 / f"validation_report_{시각}.xlsx"

    wb = Workbook()

    # ── 시트 1: 전체 오류 목록 ──
    ws1 = wb.active
    ws1.title = "오류 목록"
    ws1.append(["유형", "행", "컬럼", "값", "내용"])
    헤더_스타일(ws1)

    # 유형별 색상 (공통 색상 + 추가)
    유형색상 = {
        "필수컬럼": "FF0000",
        "빈값"    : "FF6600",
        "허용값"  : "FFC000",
        "날짜형식": "7030A0",
        "중복"    : "4472C4",
        "패턴"    : "00B0F0",
        "범위"    : "FF6600",
        "공백"    : "70AD47",
    }

    for 오류 in 오류목록:
        ws1.append([오류["유형"], str(오류["행"]), 오류["컬럼"], 오류["값"], 오류["내용"]])
        현재행 = ws1.max_row
        배경색 = 유형색상.get(오류["유형"], "BFBFBF")
        ws1.cell(현재행, 1).fill = PatternFill("solid", fgColor=배경색)
        ws1.cell(현재행, 1).font = Font(color="FFFFFF", bold=True)

    # ── 시트 2: 유형별 요약 ──
    ws2 = wb.create_sheet(title="요약")
    ws2.append(["검사 유형", "오류 건수", "결과"])
    헤더_스타일(ws2)

    검사유형 = ["필수컬럼", "빈값", "허용값", "날짜형식", "중복", "패턴", "범위", "공백"]

    for 유형 in 검사유형:
        건수 = len([o for o in 오류목록 if o["유형"] == 유형])
        결과 = "✅ 통과" if 건수 == 0 else f"❌ {건수}건"
        ws2.append([유형, 건수, 결과])
        현재행 = ws2.max_row

        배경색 = 색상_가져오기("해결") if 건수 == 0 else 색상_가져오기("Critical")
        행_색상(ws2, 현재행, 배경색)

    # 공통 모듈로 열 너비 조정
    열너비_조정(ws1)
    열너비_조정(ws2)

    wb.save(저장명)
    Latest_복사(저장명, "validation_report")
    return 저장명

def 리포트_저장_JSON(오류목록, 파일명):
    """
    검증 결과를 JSON 파일로 저장 (자동화 연동용)
    """
    결과폴더 = 결과폴더_생성(기준경로)
    시각   = 타임스탬프()
    저장명 = 결과폴더 / f"validation_report_{시각}.json"

    검사유형 = ["필수컬럼", "빈값", "허용값", "날짜형식", "중복", "패턴", "범위", "공백"]
    유형별집계 = {유형: len([o for o in 오류목록 if o["유형"] == 유형])
                  for 유형 in 검사유형}

    데이터 = {
        "검증_일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "원본_파일": str(파일명),
        "총_오류_수": len(오류목록),
        "유형별_집계": 유형별집계,
        "오류_상세": 오류목록
    }

    JSON_쓰기(저장명, 데이터)
    Latest_복사(저장명, "validation_report")
    return 저장명


def 리포트_저장_HTML(오류목록, 파일명):
    """
    검증 결과를 HTML 리포트로 저장 (웹 브라우저용)
    """
    결과폴더 = 결과폴더_생성(기준경로)
    시각   = 타임스탬프()
    저장명 = 결과폴더 / f"validation_report_{시각}.html"

    검사유형 = ["필수컬럼", "빈값", "허용값", "날짜형식", "중복", "패턴", "범위", "공백"]

    # 요약 섹션
    총오류 = len(오류목록)
    상태표시 = "✅ 데이터 무결성 통과" if 총오류 == 0 else f"[WARN] {총오류}건의 문제 발견"

    요약 = f"""
    <div class="summary">
        <h2>📊 검증 결과</h2>
        <p><strong>검증 일시:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        <p><strong>원본 파일:</strong> {파일명}</p>
        <p><strong>상태:</strong> {상태표시}</p>
    </div>
    """

    # 유형별 요약 테이블
    유형요약 = "<h2>유형별 요약</h2><table><tr><th>검사 유형</th><th>오류 건수</th><th>결과</th></tr>"
    for 유형 in 검사유형:
        건수 = len([o for o in 오류목록 if o["유형"] == 유형])
        결과 = '<span class="해결">✅ 통과</span>' if 건수 == 0 else f'<span class="미해결">❌ {건수}건</span>'
        유형요약 += f"<tr><td>{유형}</td><td>{건수}</td><td>{결과}</td></tr>"
    유형요약 += "</table>"

    # 오류 상세 테이블
    오류상세 = "<h2>🐛 오류 상세</h2>"
    if 오류목록:
        오류상세 += "<table><tr><th>유형</th><th>행</th><th>컬럼</th><th>값</th><th>내용</th></tr>"
        for 오류 in 오류목록:
            오류상세 += f"<tr><td><strong>{오류['유형']}</strong></td><td>{오류['행']}</td><td>{오류['컬럼']}</td><td>{오류['값']}</td><td>{오류['내용']}</td></tr>"
        오류상세 += "</table>"
    else:
        오류상세 += "<p>오류가 없습니다! 🎉</p>"

    본문 = 요약 + 유형요약 + 오류상세

    HTML_쓰기(저장명, "QA 데이터 검증 리포트", 본문)
    Latest_복사(저장명, "validation_report")
    return 저장명

# ────────────────────────────────────────
# ⑫ 전체 검증 실행 (메인 함수)
# ────────────────────────────────────────
# 기존 검증_실행 함수를 통째로 이거로 교체

def 검증_실행(파일명, 옵션="--xlsx"):
    """
    CSV/XLSX 파일을 읽어서 전체 검증 실행 + 다중 형식 저장

    매개변수:
    - 파일명: 검사할 파일 이름
    - 옵션  : --xlsx / --json / --html / --all (기본 --xlsx)

    반환값:
    - 딕셔너리: {"오류목록": [...], "리포트": [경로목록]}
    """
    # 1. 규칙 파일 로딩
    규칙 = 규칙_로딩(파일명)

    # 2. 파일 읽기
    df = 데이터_읽기(파일명)
    if df is None:
        return None

    # 3. 전체 검사 실행
    전체오류 = []
    전체오류 += 필수컬럼_검사(df, 규칙)
    전체오류 += 빈값_검사(df, 규칙)
    전체오류 += 허용값_검사(df, 규칙)
    전체오류 += 날짜형식_검사(df, 규칙)
    전체오류 += 중복_검사(df, 규칙)
    전체오류 += 패턴_검사(df, 규칙)
    전체오류 += 숫자범위_검사(df, 규칙)
    전체오류 += 공백_검사(df)

    # 4. 터미널에 결과 출력
    검사유형 = ["필수컬럼", "빈값", "허용값", "날짜형식", "중복", "패턴", "범위", "공백"]
    for 유형 in 검사유형:
        건수 = len([o for o in 전체오류 if o["유형"] == 유형])
        if 건수 == 0:
            print(f"[PASS] {유형:8s}")
        else:
            print(f"[FAIL] {유형:8s} : {건수}건")
            for 오류 in 전체오류:
                if 오류["유형"] == 유형:
                    print(f"   - {오류['내용']}")

    # 5. 옵션에 따라 저장
    리포트목록 = []

    if 옵션 in ["--xlsx", "--all"] or not 옵션.startswith("--"):
        경로 = 리포트_저장(전체오류, 파일명)
        print(f"[XLSX] 저장 완료: {경로}")
        리포트목록.append(경로)

    if 옵션 in ["--json", "--all"]:
        경로 = 리포트_저장_JSON(전체오류, 파일명)
        print(f"[JSON] 저장 완료: {경로}")
        리포트목록.append(경로)

    if 옵션 in ["--html", "--all"]:
        경로 = 리포트_저장_HTML(전체오류, 파일명)
        print(f"[HTML] 저장 완료: {경로}")
        리포트목록.append(경로)

    return {
        "오류목록": 전체오류,
        "리포트": 리포트목록
    }