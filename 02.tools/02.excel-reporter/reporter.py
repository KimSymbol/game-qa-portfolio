# 역할: CSV 버그 데이터를 읽어 전문적인 엑셀 리포트를 생성하는 모듈
# 다른 모듈(main.py)에서 import해서 사용
# 입력 CSV 형식: 버그ID,제목,심각도,우선순위,플랫폼,버전,상태,발견자,발견일,해결일,재현율

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# reporter.py 가 있는 폴더를 기준 경로로 설정
기준경로 = Path(__file__).parent

# ── 심각도/상태/우선순위별 색상 정의 ──
색상 = {
    # 심각도
    "Critical": "FF0000",   # 빨강
    "High"    : "FF6600",   # 주황
    "Medium"  : "FFC000",   # 노랑
    "Low"     : "70AD47",   # 초록
    # 상태
    "해결"    : "70AD47",   # 초록
    "진행중"  : "FFC000",   # 노랑
    "미해결"  : "FF0000",   # 빨강
    # 공통
    "헤더"    : "4472C4",   # 파랑
}


# ────────────────────────────────────────
# ① CSV 읽기
# ────────────────────────────────────────
def 데이터_읽기(파일명):
    """
    CSV 파일을 읽어서 DataFrame으로 반환

    매개변수:
    - 파일명: 읽을 CSV 파일 이름
              기준경로(reporter.py 위치) 기준으로 찾음

    반환값:
    - DataFrame: 버그 데이터
    - None: 파일이 없을 때

    필수 컬럼:
    버그ID, 제목, 심각도, 우선순위, 플랫폼, 버전, 상태, 발견자, 발견일

    선택 컬럼:
    해결일 (미해결/진행중이면 비워도 됨)
    재현율 (없으면 빈 값으로 처리)
    """
    경로 = 기준경로 / 파일명
    if not 경로.exists():
        print("❌ 파일이 없어요:", 경로)
        return None

    df = pd.read_csv(경로, encoding="utf-8")

    # 해결일 - 미해결/진행중이면 비어있을 수 있음
    # NaN → 빈 문자열로 변환 (엑셀에 "nan" 방지)
    if "해결일" in df.columns:
        df["해결일"] = df["해결일"].fillna("")

    # 재현율 - 선택 컬럼
    # 컬럼 자체가 없으면 빈 값으로 추가
    # 있어도 NaN이면 빈 문자열로 변환
    if "재현율" not in df.columns:
        df["재현율"] = ""
    else:
        df["재현율"] = df["재현율"].fillna("")

    print(f"✅ 데이터 로딩 완료: {len(df)}건")
    return df


# ────────────────────────────────────────
# ② 헤더 스타일 적용
# ────────────────────────────────────────
def 헤더_스타일(ws, 행번호=1):
    """
    지정한 행을 헤더 스타일로 꾸밈
    배경: 파란색 / 글씨: 흰색 굵게 / 정렬: 가운데

    매개변수:
    - ws    : openpyxl 워크시트 객체
    - 행번호: 스타일 적용할 행 번호 (기본값 1)
    """
    for 셀 in ws[행번호]:
        셀.fill      = PatternFill("solid", fgColor=색상["헤더"])
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")


# ────────────────────────────────────────
# ③ 열 너비 자동 조정
# ────────────────────────────────────────
def 열너비_조정(ws):
    """
    워크시트의 각 열 너비를 내용 길이에 맞게 자동 조정

    매개변수:
    - ws: openpyxl 워크시트 객체

    동작:
    - get_column_letter 로 병합 셀 에러 방지
    - try/except 로 병합 셀 건너뜀
    - 여백 4 추가
    """
    for i, 열 in enumerate(ws.columns, 1):
        최대길이 = 0
        열이름  = get_column_letter(i)
        for 셀 in 열:
            try:
                if 셀.value:
                    최대길이 = max(최대길이, len(str(셀.value)))
            except:
                pass
        ws.column_dimensions[열이름].width = 최대길이 + 4


# ────────────────────────────────────────
# ④ 시트 1: 전체 버그 목록
# ────────────────────────────────────────
def 시트_전체목록(wb, df):
    """
    전체 버그 데이터를 심각도별 색상으로 표시하는 시트

    색상 규칙:
    - Critical 행 → 빨간 배경
    - High 행     → 주황 배경
    - Medium 행   → 노란 배경
    - Low 행      → 초록 배경
    """
    ws = wb.active
    ws.title = "전체 버그 목록"

    # CSV 헤더를 그대로 첫 번째 행으로 사용
    ws.append(list(df.columns))
    헤더_스타일(ws)

    # 데이터 행 추가 + 심각도별 색상 적용
    for _, 행 in df.iterrows():
        ws.append(list(행))
        현재행 = ws.max_row
        심각도 = 행["심각도"]
        if 심각도 in 색상:
            for 셀 in ws[현재행]:
                셀.fill = PatternFill("solid", fgColor=색상[심각도])

    열너비_조정(ws)
    return ws


# ────────────────────────────────────────
# ⑤ 시트 2: 담당자별 현황
# ────────────────────────────────────────
def 시트_담당자별(wb, df):
    """
    발견자별 버그 상태 현황 시트
    차트 대신 색상 테이블로 표시
    - 해결  → 초록
    - 진행중 → 노랑
    - 미해결 → 빨강
    - 합계   → 파랑
    """
    ws = wb.create_sheet(title="발견자별 현황")

    # 발견자 × 상태 교차 집계 + 합계 컬럼
    담당자통계 = df.groupby(["발견자", "상태"]).size().unstack(fill_value=0)
    담당자통계["합계"] = 담당자통계.sum(axis=1)

    # 헤더
    ws.append(["발견자"] + list(담당자통계.columns))
    헤더_스타일(ws)

    # 상태별 색상 매핑
    상태색상 = {
        "해결"  : "70AD47",   # 초록
        "진행중": "FFC000",   # 노랑
        "미해결": "FF0000",   # 빨강
        "합계"  : "4472C4",   # 파랑
    }

    for 발견자, 행 in 담당자통계.iterrows():
        ws.append([발견자] + list(행))
        현재행 = ws.max_row

        # 발견자 셀 스타일
        ws.cell(현재행, 1).font      = Font(bold=True)
        ws.cell(현재행, 1).alignment = Alignment(horizontal="center")

        # 상태별 셀 색상 적용
        for col_idx, 컬럼명 in enumerate(담당자통계.columns, 2):
            셀 = ws.cell(현재행, col_idx)
            셀.alignment = Alignment(horizontal="center")
            if 컬럼명 in 상태색상:
                셀.fill = PatternFill("solid", fgColor=상태색상[컬럼명])
                셀.font = Font(color="FFFFFF", bold=True)

    열너비_조정(ws)
    return ws


# ────────────────────────────────────────
# ⑥ 시트 3: 플랫폼별 현황
# ────────────────────────────────────────
def 시트_플랫폼별(wb, df):
    """
    플랫폼별 버그 심각도 현황 시트
    차트 대신 색상 테이블로 표시
    - Critical → 빨강
    - High     → 주황
    - Medium   → 노랑
    - Low      → 초록
    - 합계     → 파랑
    """
    ws = wb.create_sheet(title="플랫폼별 현황")

    # 플랫폼 × 심각도 교차 집계 + 합계 컬럼
    플랫폼통계 = df.groupby(["플랫폼", "심각도"]).size().unstack(fill_value=0)
    플랫폼통계["합계"] = 플랫폼통계.sum(axis=1)

    # 헤더
    ws.append(["플랫폼"] + list(플랫폼통계.columns))
    헤더_스타일(ws)

    for 플랫폼, 행 in 플랫폼통계.iterrows():
        ws.append([플랫폼] + list(행))
        현재행 = ws.max_row

        # 플랫폼 셀 스타일
        ws.cell(현재행, 1).font      = Font(bold=True)
        ws.cell(현재행, 1).alignment = Alignment(horizontal="center")

        # 심각도별 셀 색상 적용
        for col_idx, 컬럼명 in enumerate(플랫폼통계.columns, 2):
            셀 = ws.cell(현재행, col_idx)
            셀.alignment = Alignment(horizontal="center")
            배경색 = 색상.get(컬럼명, "BFBFBF")  # 없는 값은 회색
            셀.fill = PatternFill("solid", fgColor=배경색)
            셀.font = Font(color="FFFFFF", bold=True)

    열너비_조정(ws)
    return ws


# ────────────────────────────────────────
# ⑦ 시트 4: 심각도 + 우선순위 통계
# ────────────────────────────────────────
def 시트_심각도별(wb, df):
    """
    심각도 및 우선순위별 버그 건수를 색상 구분해서 표시하는 시트
    """
    ws = wb.create_sheet(title="심각도·우선순위")

    # 심각도별 집계
    ws.append(["심각도", "건수"])
    헤더_스타일(ws)

    심각도통계 = df["심각도"].value_counts()
    for 심각도, 건수 in 심각도통계.items():
        ws.append([심각도, 건수])
        현재행 = ws.max_row
        if 심각도 in 색상:
            for 셀 in ws[현재행]:
                셀.fill      = PatternFill("solid", fgColor=색상[심각도])
                셀.font      = Font(color="FFFFFF", bold=True)
                셀.alignment = Alignment(horizontal="center")

    # 빈 행 구분
    ws.append(["", ""])

    # 우선순위별 집계
    ws.append(["우선순위", "건수"])
    헤더_스타일(ws, ws.max_row)

    우선순위색상 = {"High": "FF0000", "Medium": "FFC000", "Low": "70AD47"}
    우선순위통계 = df["우선순위"].value_counts()
    for 우선순위, 건수 in 우선순위통계.items():
        ws.append([우선순위, 건수])
        현재행 = ws.max_row
        if 우선순위 in 우선순위색상:
            for 셀 in ws[현재행]:
                셀.fill      = PatternFill("solid", fgColor=우선순위색상[우선순위])
                셀.font      = Font(color="FFFFFF", bold=True)
                셀.alignment = Alignment(horizontal="center")

    열너비_조정(ws)
    return ws


# ────────────────────────────────────────
# ⑧ 시트 5: 요약 대시보드
# ────────────────────────────────────────
def 시트_대시보드(wb, df):
    """
    핵심 지표를 한눈에 보여주는 대시보드 시트

    표시 항목:
    - 총 버그 수 / 해결 / 진행중 / 미해결 / 해결률
    - 심각도별 건수
    - 우선순위별 건수
    - 플랫폼별 건수
    """
    ws = wb.create_sheet(title="요약 대시보드")

    총버그 = len(df)
    상태목록 = df["상태"].unique()
    해결수  = len(df[df["상태"] == "해결"]) if "해결" in 상태목록 else 0
    해결률  = round(해결수 / 총버그 * 100, 1) if 총버그 > 0 else 0

    기본데이터 = [
        ["📊 QA 버그 리포트 요약", ""],
        ["", ""],
        ["항목",      "수치"],
        ["총 버그 수", 총버그],
        ["해결률",     f"{해결률}%"],
        ["", ""],
    ]

    for 행 in 기본데이터:
        ws.append(행)

    # 상태별
    ws.append(["상태별", "건수"])
    헤더_스타일(ws, ws.max_row)
    상태색상 = {"해결": "70AD47", "진행중": "FFC000", "미해결": "FF0000"}
    for 상태 in 상태목록:
        ws.append([상태, len(df[df["상태"] == 상태])])
        현재행 = ws.max_row
        배경색 = 상태색상.get(상태, "BFBFBF")
        for 셀 in ws[현재행]:
            셀.fill      = PatternFill("solid", fgColor=배경색)
            셀.font      = Font(color="FFFFFF", bold=True)
            셀.alignment = Alignment(horizontal="center")

    ws.append(["", ""])

    # 심각도별
    ws.append(["심각도별", "건수"])
    헤더_스타일(ws, ws.max_row)
    for 심각도 in df["심각도"].unique():
        ws.append([심각도, len(df[df["심각도"] == 심각도])])
        현재행 = ws.max_row
        배경색 = 색상.get(심각도, "BFBFBF")
        for 셀 in ws[현재행]:
            셀.fill      = PatternFill("solid", fgColor=배경색)
            셀.font      = Font(color="FFFFFF", bold=True)
            셀.alignment = Alignment(horizontal="center")

    ws.append(["", ""])

    # 우선순위별
    ws.append(["우선순위별", "건수"])
    헤더_스타일(ws, ws.max_row)
    우선순위색상 = {"High": "FF0000", "Medium": "FFC000", "Low": "70AD47"}
    for 우선순위 in df["우선순위"].unique():
        ws.append([우선순위, len(df[df["우선순위"] == 우선순위])])
        현재행 = ws.max_row
        배경색 = 우선순위색상.get(우선순위, "BFBFBF")
        for 셀 in ws[현재행]:
            셀.fill      = PatternFill("solid", fgColor=배경색)
            셀.font      = Font(color="FFFFFF", bold=True)
            셀.alignment = Alignment(horizontal="center")

    ws.append(["", ""])

    # 플랫폼별
    ws.append(["플랫폼별", "건수"])
    헤더_스타일(ws, ws.max_row)
    for 플랫폼 in df["플랫폼"].unique():
        ws.append([플랫폼, len(df[df["플랫폼"] == 플랫폼])])
        현재행 = ws.max_row
        for 셀 in ws[현재행]:
            셀.alignment = Alignment(horizontal="center")

    # 제목 스타일
    ws["A1"].font      = Font(size=16, bold=True, color="4472C4")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    # 해결률 강조
    ws["B5"].font = Font(bold=True, color="70AD47", size=12)

    # ← 병합 셀 때문에 열너비_조정이 B열을 못 잡음
    # 직접 너비 지정으로 해결
    ws.column_dimensions["A"].width = 20   # A열 너비
    ws.column_dimensions["B"].width = 10   # B열 너비

    return ws


# ────────────────────────────────────────
# ⑨ 전체 리포트 생성
# ────────────────────────────────────────
def 리포트_생성(df):
    """
    5시트 엑셀 리포트 생성 및 저장

    매개변수:
    - df: 버그 데이터 DataFrame

    반환값:
    - 저장된 파일 경로 (Path 객체)

    저장 위치: 결과/report_YYYY-MM-DD.xlsx

    시트 구성:
    1. 전체 버그 목록  - 심각도별 색상 구분
    2. 발견자별 현황   - 교차 집계 + 차트
    3. 플랫폼별 현황   - 교차 집계 + 차트
    4. 심각도·우선순위 - 집계
    5. 요약 대시보드   - 핵심 지표
    """
    결과폴더 = 기준경로 / "결과"
    결과폴더.mkdir(exist_ok=True)

    오늘   = datetime.now().strftime("%Y-%m-%d")
    파일명 = 결과폴더 / f"report_{오늘}.xlsx"

    wb = Workbook()

    시트_전체목록(wb, df)
    시트_담당자별(wb, df)
    시트_플랫폼별(wb, df)
    시트_심각도별(wb, df)
    시트_대시보드(wb, df)

    wb.save(파일명)
    return 파일명