# 역할: CSV/XLSX 버그 데이터를 읽어 엑셀 리포트를 생성하는 모듈
# 공통 모듈: common.file_io, common.excel_style 사용
#
# 입력 형식: 버그ID,제목,심각도,우선순위,플랫폼,버전,상태,발견자,발견일,해결일,재현율

import sys
from pathlib import Path

# 02.tools 폴더를 Python 경로에 추가 (common 모듈 import 위해)
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

# 공통 모듈
from common.file_io     import (
    파일_읽기, 결과폴더_생성, NaN_처리, 타임스탬프, Latest_복사,
    JSON_쓰기, HTML_쓰기,PDF_쓰기
)
from common.excel_style import (
    헤더_스타일, 행_색상, 색상_가져오기,
    열너비_조정, 색상, 우선순위_색상
)

# reporter.py 가 있는 폴더를 기준 경로로 설정
기준경로 = Path(__file__).parent


# ────────────────────────────────────────
# ① 데이터 읽기 (csv / xlsx 자동 판단)
# ────────────────────────────────────────
def 데이터_읽기(파일명):
    """
    CSV 또는 XLSX 파일을 읽어서 DataFrame 반환

    매개변수:
    - 파일명: 읽을 파일 이름 또는 경로

    반환값:
    - DataFrame
    - None: 파일이 없거나 지원하지 않는 형식

    동작:
    - common.file_io.파일_읽기() 사용
    - 해결일/재현율 NaN → 빈 문자열 처리
    """
    df = 파일_읽기(파일명, 기준경로)
    if df is None:
        return None

    # 해결일 / 재현율 NaN 처리
    if "해결일" in df.columns:
        df["해결일"] = df["해결일"].fillna("")
    if "재현율" not in df.columns:
        df["재현율"] = ""
    else:
        df["재현율"] = df["재현율"].fillna("")

    return df


# ────────────────────────────────────────
# ② 시트 1: 전체 버그 목록
# ────────────────────────────────────────
def 시트_전체목록(wb, df):
    """
    전체 버그 데이터를 심각도별 색상으로 표시하는 시트
    """
    ws = wb.active
    ws.title = "전체 버그 목록"

    ws.append(list(df.columns))
    헤더_스타일(ws)

    for _, 행 in df.iterrows():
        ws.append(list(행))
        현재행 = ws.max_row
        심각도 = 행["심각도"]
        배경색 = 색상_가져오기(심각도)
        행_색상(ws, 현재행, 배경색)

    열너비_조정(ws)
    return ws


# ────────────────────────────────────────
# ③ 시트 2: 발견자별 현황
# ────────────────────────────────────────
def 시트_발견자별(wb, df):
    """
    발견자별 버그 상태 현황 시트 (색상 테이블)
    """
    ws = wb.create_sheet(title="발견자별 현황")

    담당자통계 = df.groupby(["발견자", "상태"]).size().unstack(fill_value=0)
    담당자통계["합계"] = 담당자통계.sum(axis=1)

    ws.append(["발견자"] + list(담당자통계.columns))
    헤더_스타일(ws)

    상태색상 = {
        "해결"  : 색상_가져오기("해결"),
        "진행중": 색상_가져오기("진행중"),
        "미해결": 색상_가져오기("미해결"),
        "합계"  : 색상_가져오기("헤더"),
    }

    for 발견자, 행 in 담당자통계.iterrows():
        ws.append([발견자] + list(행))
        현재행 = ws.max_row

        ws.cell(현재행, 1).font      = Font(bold=True)
        ws.cell(현재행, 1).alignment = Alignment(horizontal="center")

        for col_idx, 컬럼명 in enumerate(담당자통계.columns, 2):
            셀 = ws.cell(현재행, col_idx)
            셀.alignment = Alignment(horizontal="center")
            if 컬럼명 in 상태색상:
                셀.fill = PatternFill("solid", fgColor=상태색상[컬럼명])
                셀.font = Font(color="FFFFFF", bold=True)

    열너비_조정(ws)
    return ws


# ────────────────────────────────────────
# ④ 시트 3: 플랫폼별 현황
# ────────────────────────────────────────
def 시트_플랫폼별(wb, df):
    """
    플랫폼별 버그 심각도 현황 시트 (색상 테이블)
    """
    ws = wb.create_sheet(title="플랫폼별 현황")

    플랫폼통계 = df.groupby(["플랫폼", "심각도"]).size().unstack(fill_value=0)
    플랫폼통계["합계"] = 플랫폼통계.sum(axis=1)

    ws.append(["플랫폼"] + list(플랫폼통계.columns))
    헤더_스타일(ws)

    for 플랫폼, 행 in 플랫폼통계.iterrows():
        ws.append([플랫폼] + list(행))
        현재행 = ws.max_row

        ws.cell(현재행, 1).font      = Font(bold=True)
        ws.cell(현재행, 1).alignment = Alignment(horizontal="center")

        for col_idx, 컬럼명 in enumerate(플랫폼통계.columns, 2):
            셀 = ws.cell(현재행, col_idx)
            셀.alignment = Alignment(horizontal="center")
            배경색 = 색상_가져오기(컬럼명)
            셀.fill = PatternFill("solid", fgColor=배경색)
            셀.font = Font(color="FFFFFF", bold=True)

    열너비_조정(ws)
    return ws


# ────────────────────────────────────────
# ⑤ 시트 4: 심각도·우선순위 통계
# ────────────────────────────────────────
def 시트_심각도별(wb, df):
    """
    심각도 및 우선순위별 버그 건수
    """
    ws = wb.create_sheet(title="심각도·우선순위")

    ws.append(["심각도", "건수"])
    헤더_스타일(ws)

    for 심각도 in df["심각도"].unique():
        ws.append([심각도, len(df[df["심각도"] == 심각도])])
        현재행 = ws.max_row
        배경색 = 색상_가져오기(심각도)
        행_색상(ws, 현재행, 배경색)

    ws.append(["", ""])

    ws.append(["우선순위", "건수"])
    헤더_스타일(ws, ws.max_row)

    for 우선순위 in df["우선순위"].unique():
        ws.append([우선순위, len(df[df["우선순위"] == 우선순위])])
        현재행 = ws.max_row
        배경색 = 우선순위_색상.get(우선순위, 색상_가져오기("기본"))
        행_색상(ws, 현재행, 배경색)

    열너비_조정(ws)
    return ws


# ────────────────────────────────────────
# ⑥ 시트 5: 요약 대시보드
# ────────────────────────────────────────
def 시트_대시보드(wb, df):
    """
    핵심 지표를 한눈에 보여주는 대시보드
    """
    ws = wb.create_sheet(title="요약 대시보드")

    총버그 = len(df)
    상태목록 = df["상태"].unique()
    해결수   = len(df[df["상태"] == "해결"]) if "해결" in 상태목록 else 0
    해결률   = round(해결수 / 총버그 * 100, 1) if 총버그 > 0 else 0

    기본데이터 = [
        ["QA 버그 리포트 요약", ""],
        ["", ""],
        ["항목",      "수치"],
        ["총 버그 수", 총버그],
        ["해결률",     f"{해결률}%"],
        ["", ""],
    ]

    for 행 in 기본데이터:
        ws.append(행)

    # 상태별
    ws.append(["상태별", "건수"])
    헤더_스타일(ws, ws.max_row)
    for 상태 in 상태목록:
        ws.append([상태, len(df[df["상태"] == 상태])])
        현재행 = ws.max_row
        배경색 = 색상_가져오기(상태)
        행_색상(ws, 현재행, 배경색)

    ws.append(["", ""])

    # 심각도별
    ws.append(["심각도별", "건수"])
    헤더_스타일(ws, ws.max_row)
    for 심각도 in df["심각도"].unique():
        ws.append([심각도, len(df[df["심각도"] == 심각도])])
        현재행 = ws.max_row
        배경색 = 색상_가져오기(심각도)
        행_색상(ws, 현재행, 배경색)

    ws.append(["", ""])

    # 우선순위별
    ws.append(["우선순위별", "건수"])
    헤더_스타일(ws, ws.max_row)
    for 우선순위 in df["우선순위"].unique():
        ws.append([우선순위, len(df[df["우선순위"] == 우선순위])])
        현재행 = ws.max_row
        배경색 = 우선순위_색상.get(우선순위, 색상_가져오기("기본"))
        행_색상(ws, 현재행, 배경색)

    ws.append(["", ""])

    # 플랫폼별
    ws.append(["플랫폼별", "건수"])
    헤더_스타일(ws, ws.max_row)
    for 플랫폼 in df["플랫폼"].unique():
        ws.append([플랫폼, len(df[df["플랫폼"] == 플랫폼])])
        현재행 = ws.max_row
        for 셀 in ws[현재행]:
            셀.alignment = Alignment(horizontal="center")

    # 제목 스타일
    ws["A1"].font      = Font(size=16, bold=True, color="4472C4")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    # 해결률 강조
    ws["B5"].font = Font(bold=True, color="70AD47", size=12)

    # 병합 셀 때문에 열너비_조정 못 씀 → 직접 지정
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10

    return ws


# ────────────────────────────────────────
# ⑦ 전체 리포트 생성
# ────────────────────────────────────────
def 리포트_생성(df):
    """5시트 엑셀 리포트 생성"""
    결과폴더 = 결과폴더_생성(기준경로)
    시각 = 타임스탬프()
    파일명 = 결과폴더 / f"report_{시각}.xlsx"

    wb = Workbook()
    시트_전체목록(wb, df)
    시트_발견자별(wb, df)
    시트_플랫폼별(wb, df)
    시트_심각도별(wb, df)
    시트_대시보드(wb, df)

    wb.save(파일명)
    Latest_복사(파일명, "report")  # ← latest 복사
    return 파일명


def 리포트_생성_JSON(df):
    """JSON 리포트 생성"""
    결과폴더 = 결과폴더_생성(기준경로)
    시각 = 타임스탬프()
    파일명 = 결과폴더 / f"report_{시각}.json"

    데이터 = {
        "생성일": 시각,
        "통계": {
            "총_버그_수": len(df),
            "해결": int(len(df[df["상태"] == "해결"])),
            "진행중": int(len(df[df["상태"] == "진행중"])),
            "미해결": int(len(df[df["상태"] == "미해결"])),
            "심각도별": df["심각도"].value_counts().to_dict(),
            "우선순위별": df["우선순위"].value_counts().to_dict(),
            "플랫폼별": df["플랫폼"].value_counts().to_dict(),
        },
        "상세": df.to_dict(orient="records")
    }

    JSON_쓰기(파일명, 데이터)
    Latest_복사(파일명, "report")
    return 파일명


def 리포트_생성_HTML(df):
    """HTML 리포트 생성"""
    결과폴더 = 결과폴더_생성(기준경로)
    시각 = 타임스탬프()
    파일명 = 결과폴더 / f"report_{시각}.html"

    # 기존 HTML 생성 로직 그대로
    총버그 = len(df)
    해결   = len(df[df["상태"] == "해결"])
    해결률 = round(해결 / 총버그 * 100, 1) if 총버그 > 0 else 0

    요약 = f"""
    <div class="summary">
        <h2>요약</h2>
        <p><strong>생성일:</strong> {시각}</p>
        <p><strong>총 버그 수:</strong> {총버그}건</p>
        <p><strong>해결률:</strong> <span style="color:#70AD47; font-weight:bold;">{해결률}%</span></p>
    </div>
    """

    심각도통계 = "<h2>🎯 심각도별</h2><table><tr><th>심각도</th><th>건수</th></tr>"
    for 심각도, 건수 in df["심각도"].value_counts().items():
        심각도통계 += f'<tr><td class="{심각도}">{심각도}</td><td>{건수}</td></tr>'
    심각도통계 += "</table>"

    상태통계 = "<h2>📋 상태별</h2><table><tr><th>상태</th><th>건수</th></tr>"
    for 상태, 건수 in df["상태"].value_counts().items():
        상태통계 += f'<tr><td class="{상태}">{상태}</td><td>{건수}</td></tr>'
    상태통계 += "</table>"

    버그목록 = "<h2>🐛 전체 버그 목록</h2>"
    버그목록 += df.to_html(index=False, classes="bug-table", escape=False)

    본문 = 요약 + 심각도통계 + 상태통계 + 버그목록

    HTML_쓰기(파일명, "QA 버그 리포트", 본문)
    Latest_복사(파일명, "report")
    return 파일명


def 리포트_생성_PDF(df):
    """PDF 리포트 생성"""
    결과폴더 = 결과폴더_생성(기준경로)
    시각 = 타임스탬프()
    파일명 = 결과폴더 / f"report_{시각}.pdf"

    총버그 = len(df)
    해결   = len(df[df["상태"] == "해결"])
    해결률 = round(해결 / 총버그 * 100, 1) if 총버그 > 0 else 0

    요약 = {
        "생성일": 시각,
        "총 버그 수": f"{총버그}건",
        "해결": f"{해결}건",
        "해결률": f"{해결률}%",
    }

    PDF_쓰기(파일명, "QA 버그 리포트", df, 요약)
    Latest_복사(파일명, "report")
    return 파일명