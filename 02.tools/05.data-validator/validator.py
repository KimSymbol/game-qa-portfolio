# 역할: CSV 파일의 데이터 무결성을 자동으로 검사하는 모듈
# 검사 규칙은 rules/ 폴더의 JSON 파일로 관리
# 규칙 파일 없으면 기본 규칙(빈 값, 중복, 공백)으로 검사
#
# 검사 항목:
#   - 필수 컬럼 존재 여부
#   - 필수 값 누락 (빈 값)
#   - 허용되지 않는 값 (오타)
#   - 날짜 형식 오류
#   - 중복 데이터
#   - 패턴 오류 (정규표현식)
#   - 공백 오류

import re
import json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# validator.py 가 있는 폴더를 기준 경로로 설정
기준경로 = Path(__file__).parent


# ────────────────────────────────────────
# ① 규칙 파일 로딩
# ────────────────────────────────────────
def 규칙_로딩(파일명):
    """
    CSV 파일 이름에 맞는 규칙 JSON 파일을 자동으로 찾아서 로딩

    매개변수:
    - 파일명: 검사할 CSV 파일 이름

    반환값:
    - 딕셔너리: 규칙 데이터
    - 기본 규칙: 매칭되는 규칙 파일 없을 때

    규칙 파일 매핑:
    - 파일명에 "bugs" 포함     → rules/bugs.json
    - 파일명에 "testcases" 포함 → rules/testcases.json
    - 매칭 안 되면             → rules/default.json
    - default.json 도 없으면   → 내장 기본 규칙
    """
    규칙폴더 = 기준경로 / "rules"
    파일명소문자 = str(파일명).lower()

    # 파일명으로 규칙 파일 자동 매핑
    규칙파일 = None
    if "bugs" in 파일명소문자:
        규칙파일 = 규칙폴더 / "bugs.json"
    elif "testcases" in 파일명소문자:
        규칙파일 = 규칙폴더 / "testcases.json"

    # 매핑 안 되면 default.json
    if 규칙파일 is None or not 규칙파일.exists():
        규칙파일 = 규칙폴더 / "default.json"

    # default.json 도 없으면 내장 기본 규칙
    if not 규칙파일.exists():
        print("⚠️  규칙 파일 없음 → 기본 규칙으로 검사")
        return {
            "설명": "내장 기본 규칙",
            "필수_컬럼": [],
            "컬럼_규칙": {}
        }

    with open(규칙파일, "r", encoding="utf-8") as f:
        규칙 = json.load(f)

    print(f"📋 규칙 파일 로딩: {규칙파일.name} ({규칙.get('설명', '')})")
    return 규칙


# ────────────────────────────────────────
# ② CSV 파일 읽기
# ────────────────────────────────────────
def 데이터_읽기(파일명):
    """
    CSV 파일을 읽어서 DataFrame으로 반환

    매개변수:
    - 파일명: 읽을 CSV 파일 이름

    반환값:
    - DataFrame
    - None: 파일이 없을 때
    """
    경로 = Path(파일명)
    if not 경로.exists():
        경로 = 기준경로 / 파일명
    if not 경로.exists():
        print("❌ 파일이 없어요:", 파일명)
        return None

    df = pd.read_csv(경로, encoding="utf-8-sig")
    print(f"✅ 데이터 로딩 완료: {len(df)}건")
    return df


# ────────────────────────────────────────
# ③ 필수 컬럼 검사
# ────────────────────────────────────────
def 필수컬럼_검사(df, 규칙):
    """
    필수 컬럼이 CSV에 존재하는지 검사

    반환값:
    - 오류 리스트: [{"유형": "필수컬럼", "내용": "..."}]
    """
    오류목록 = []
    필수컬럼 = 규칙.get("필수_컬럼", [])

    for 컬럼 in 필수컬럼:
        if 컬럼 not in df.columns:
            오류목록.append({
                "유형": "필수컬럼",
                "행": "-",
                "컬럼": 컬럼,
                "값": "-",
                "내용": f"필수 컬럼 '{컬럼}' 이 없습니다"
            })

    return 오류목록


# ────────────────────────────────────────
# ④ 빈 값 검사
# ────────────────────────────────────────
def 빈값_검사(df, 규칙):
    """
    필수 항목에 빈 값이 있는지 검사
    규칙 파일의 "필수": true 인 컬럼만 검사
    규칙 파일 없으면 모든 컬럼 검사

    반환값:
    - 오류 리스트
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    # 검사할 컬럼 결정
    if 규칙:
        # 규칙 파일 있음 → 필수 컬럼만 검사
        검사컬럼 = [컬럼 for 컬럼, 설정 in 컬럼규칙.items()
                    if 설정.get("필수", False) and 컬럼 in df.columns]
    else:
        # 규칙 파일 없음 → 모든 컬럼 검사
        검사컬럼 = list(df.columns)

    for 컬럼 in 검사컬럼:
        빈행목록 = df[df[컬럼].isna() | (df[컬럼].astype(str).str.strip() == "")]
        for idx in 빈행목록.index:
            오류목록.append({
                "유형": "빈값",
                "행": idx + 2,     # 헤더 = 1행이니까 +2
                "컬럼": 컬럼,
                "값": "(비어있음)",
                "내용": f"{idx+2}행 '{컬럼}' 값이 비어있습니다"
            })

    return 오류목록


# ────────────────────────────────────────
# ⑤ 허용값 검사
# ────────────────────────────────────────
def 허용값_검사(df, 규칙):
    """
    특정 컬럼의 값이 허용된 값 목록에 포함되는지 검사
    오타 감지에 핵심 역할

    반환값:
    - 오류 리스트
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    for 컬럼, 설정 in 컬럼규칙.items():
        허용값 = 설정.get("허용값", None)
        if 허용값 is None or 컬럼 not in df.columns:
            continue

        for idx, 행 in df.iterrows():
            값 = str(행[컬럼]).strip()
            if 값 == "" or 값 == "nan":
                continue  # 빈 값은 빈값_검사에서 처리
            if 값 not in 허용값:
                오류목록.append({
                    "유형": "허용값",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": 값,
                    "내용": f"{idx+2}행 '{컬럼}' 값 '{값}' 은 허용되지 않음 (허용값: {', '.join(허용값)})"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑥ 날짜 형식 검사
# ────────────────────────────────────────
def 날짜형식_검사(df, 규칙):
    """
    날짜 컬럼의 값이 올바른 형식인지 검사
    예: "2026-06-17" → YYYY-MM-DD 형식

    반환값:
    - 오류 리스트
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    for 컬럼, 설정 in 컬럼규칙.items():
        날짜형식 = 설정.get("날짜형식", None)
        if 날짜형식 is None or 컬럼 not in df.columns:
            continue

        for idx, 행 in df.iterrows():
            값 = str(행[컬럼]).strip()
            if 값 == "" or 값 == "nan":
                continue  # 빈 값은 빈값_검사에서 처리
            try:
                datetime.strptime(값, 날짜형식)
            except ValueError:
                오류목록.append({
                    "유형": "날짜형식",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": 값,
                    "내용": f"{idx+2}행 '{컬럼}' 값 '{값}' 은 올바른 날짜 형식이 아님 (형식: {날짜형식})"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑦ 중복 데이터 검사
# ────────────────────────────────────────
def 중복_검사(df, 규칙):
    """
    중복불가로 지정된 컬럼에 중복 값이 있는지 검사
    규칙 파일 없으면 전체 행 중복 검사

    반환값:
    - 오류 리스트
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    # 중복불가 컬럼 검사
    중복검사컬럼 = [컬럼 for 컬럼, 설정 in 컬럼규칙.items()
                    if 설정.get("중복불가", False) and 컬럼 in df.columns]

    if 중복검사컬럼:
        for 컬럼 in 중복검사컬럼:
            중복 = df[df[컬럼].duplicated(keep=False)]
            중복값목록 = 중복[컬럼].unique()
            for 값 in 중복값목록:
                중복행 = df[df[컬럼] == 값].index.tolist()
                행번호 = ", ".join([str(i+2) for i in 중복행])
                오류목록.append({
                    "유형": "중복",
                    "행": 행번호,
                    "컬럼": 컬럼,
                    "값": str(값),
                    "내용": f"'{컬럼}' 값 '{값}' 이 중복됨 ({행번호}행)"
                })
    else:
        # 규칙 없으면 전체 행 중복 검사
        중복행 = df[df.duplicated(keep=False)]
        if len(중복행) > 0:
            오류목록.append({
                "유형": "중복",
                "행": "-",
                "컬럼": "전체",
                "값": f"{len(중복행)}건",
                "내용": f"완전히 동일한 행이 {len(중복행)}건 발견됨"
            })

    return 오류목록


# ────────────────────────────────────────
# ⑧ 패턴 검사 (정규표현식)
# ────────────────────────────────────────
def 패턴_검사(df, 규칙):
    """
    정규표현식 패턴에 맞지 않는 값 검사
    예: 버그ID가 "BUG-숫자" 형식인지

    반환값:
    - 오류 리스트
    """
    오류목록 = []
    컬럼규칙 = 규칙.get("컬럼_규칙", {})

    for 컬럼, 설정 in 컬럼규칙.items():
        패턴 = 설정.get("패턴", None)
        if 패턴 is None or 컬럼 not in df.columns:
            continue

        for idx, 행 in df.iterrows():
            값 = str(행[컬럼]).strip()
            if 값 == "" or 값 == "nan":
                continue
            if not re.fullmatch(패턴, 값):
                오류목록.append({
                    "유형": "패턴",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": 값,
                    "내용": f"{idx+2}행 '{컬럼}' 값 '{값}' 이 패턴에 맞지 않음 (패턴: {패턴})"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑨ 공백 오류 검사
# ────────────────────────────────────────
def 공백_검사(df):
    """
    모든 컬럼의 값에서 앞뒤 공백 오류 검사
    규칙 파일 없어도 항상 동작하는 기본 검사

    반환값:
    - 오류 리스트
    """
    오류목록 = []

    for 컬럼 in df.columns:
        for idx, 행 in df.iterrows():
            값 = 행[컬럼]
            if pd.isna(값):
                continue
            값문자열 = str(값)
            if 값문자열 != 값문자열.strip():
                오류목록.append({
                    "유형": "공백",
                    "행": idx + 2,
                    "컬럼": 컬럼,
                    "값": repr(값문자열),
                    "내용": f"{idx+2}행 '{컬럼}' 값에 앞뒤 공백이 있음"
                })

    return 오류목록


# ────────────────────────────────────────
# ⑩ 열 너비 자동 조정
# ────────────────────────────────────────
def 열너비_조정(ws):
    """워크시트의 각 열 너비를 내용 길이에 맞게 자동 조정"""
    for i, 열 in enumerate(ws.columns, 1):
        최대길이 = 0
        열이름  = get_column_letter(i)
        for 셀 in 열:
            try:
                if 셀.value:
                    최대길이 = max(최대길이, len(str(셀.value)))
            except:
                pass
        ws.column_dimensions[열이름].width = min(최대길이 + 4, 60)


# ────────────────────────────────────────
# ⑪ 검증 리포트 엑셀 저장
# ────────────────────────────────────────
def 리포트_저장(오류목록, 파일명):
    """
    검증 결과를 엑셀 파일로 저장

    매개변수:
    - 오류목록: 전체 오류 리스트
    - 파일명  : 검사한 원본 파일 이름

    저장 위치: 결과/validation_report_YYYY-MM-DD.xlsx

    시트 구성:
    - 전체 오류 목록 (색상 구분)
    - 유형별 요약 통계
    """
    결과폴더 = 기준경로 / "결과"
    결과폴더.mkdir(exist_ok=True)

    오늘   = datetime.now().strftime("%Y-%m-%d")
    저장명 = 결과폴더 / f"validation_report_{오늘}.xlsx"

    wb = Workbook()

    # ── 시트 1: 전체 오류 목록 ──
    ws1 = wb.active
    ws1.title = "오류 목록"
    ws1.append(["유형", "행", "컬럼", "값", "내용"])

    # 헤더 스타일
    for 셀 in ws1[1]:
        셀.fill      = PatternFill("solid", fgColor="4472C4")
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")

    # 유형별 색상
    유형색상 = {
        "필수컬럼": "FF0000",
        "빈값"    : "FF6600",
        "허용값"  : "FFC000",
        "날짜형식": "7030A0",
        "중복"    : "4472C4",
        "패턴"    : "00B0F0",
        "공백"    : "70AD47",
    }

    for 오류 in 오류목록:
        ws1.append([오류["유형"], str(오류["행"]), 오류["컬럼"], 오류["값"], 오류["내용"]])
        현재행 = ws1.max_row
        색상 = 유형색상.get(오류["유형"], "BFBFBF")
        ws1.cell(현재행, 1).fill = PatternFill("solid", fgColor=색상)
        ws1.cell(현재행, 1).font = Font(color="FFFFFF", bold=True)

    # ── 시트 2: 유형별 요약 ──
    ws2 = wb.create_sheet(title="요약")
    ws2.append(["검사 유형", "오류 건수", "결과"])

    for 셀 in ws2[1]:
        셀.fill      = PatternFill("solid", fgColor="4472C4")
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")

    검사유형 = ["필수컬럼", "빈값", "허용값", "날짜형식", "중복", "패턴", "공백"]

    for 유형 in 검사유형:
        건수 = len([o for o in 오류목록 if o["유형"] == 유형])
        결과 = "✅ 통과" if 건수 == 0 else f"❌ {건수}건"
        ws2.append([유형, 건수, 결과])
        현재행 = ws2.max_row

        색상 = "70AD47" if 건수 == 0 else "FF0000"
        for 셀 in ws2[현재행]:
            셀.fill      = PatternFill("solid", fgColor=색상)
            셀.font      = Font(color="FFFFFF", bold=True)
            셀.alignment = Alignment(horizontal="center")

    # 열 너비 조정
    열너비_조정(ws1)
    열너비_조정(ws2)

    wb.save(저장명)
    return 저장명


# ────────────────────────────────────────
# ⑫ 전체 검증 실행 (메인 함수)
# ────────────────────────────────────────
def 검증_실행(파일명):
    """
    CSV 파일을 읽어서 전체 검증 실행

    매개변수:
    - 파일명: 검사할 CSV 파일 이름

    동작 순서:
    1. 규칙 파일 로딩 (자동 매핑)
    2. CSV 파일 읽기
    3. 전체 검사 실행 (7종)
    4. 터미널에 결과 출력
    5. 검증 리포트 엑셀 저장

    반환값:
    - 딕셔너리: {"오류목록": [...], "리포트": 경로}
    """
    # 1. 규칙 파일 로딩
    규칙 = 규칙_로딩(파일명)

    # 2. CSV 읽기
    df = 데이터_읽기(파일명)
    if df is None:
        return None

    # 3. 전체 검사 실행
    전체오류 = []
    전체오류 += 필수컬럼_검사(df, 규칙)
    전체오류 += 빈값_검사(df, 규칙)
    전체오류 += 허용값_검사(df, 규칙)
    전체오류 += 날짜형식_검사(df, 규칙)
    전체오류 += 중복_검사(df, 규칙)
    전체오류 += 패턴_검사(df, 규칙)
    전체오류 += 공백_검사(df)

    # 4. 터미널에 결과 출력
    검사유형 = ["필수컬럼", "빈값", "허용값", "날짜형식", "중복", "패턴", "공백"]
    for 유형 in 검사유형:
        건수 = len([o for o in 전체오류 if o["유형"] == 유형])
        if 건수 == 0:
            print(f"✅ {유형:8s} : 통과")
        else:
            print(f"❌ {유형:8s} : {건수}건")
            for 오류 in 전체오류:
                if 오류["유형"] == 유형:
                    print(f"   - {오류['내용']}")

    # 5. 리포트 저장
    리포트경로 = 리포트_저장(전체오류, 파일명)

    return {
        "오류목록": 전체오류,
        "리포트": 리포트경로
    }