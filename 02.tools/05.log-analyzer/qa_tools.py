# 역할: 로그 파일 분석에 필요한 핵심 함수 모음
# 공통 모듈: common.file_io, common.excel_style 사용
#
# 입력 형식: .txt 로그 파일
# 출력: 5시트 엑셀 리포트 + bug_history.json

import sys
import re
import json
from pathlib import Path
from datetime import datetime
from collections import Counter

# 02.tools 폴더를 Python 경로에 추가
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# 공통 모듈
from common.file_io import (
    파일_읽기, 결과폴더_생성, JSON_쓰기, HTML_쓰기,
    타임스탬프, Latest_복사
)
from common.excel_style import (
    헤더_스타일, 행_색상, 색상_가져오기, 열너비_조정, 색상
)

# qa_tools.py 가 있는 폴더를 기준 경로로 설정
기준경로 = Path(__file__).parent


# ────────────────────────────────────────
# ① 로그 파일 읽기
# ────────────────────────────────────────
def 로그_읽기(파일명):
    """
    로그 파일을 읽어서 줄 목록으로 반환
    common.file_io.파일_읽기 사용 (.txt 자동 처리)

    반환값:
    - 줄 목록 (list)
    - []: 파일이 없거나 빈 파일
    """
    줄목록 = 파일_읽기(파일명, 기준경로)
    if 줄목록 is None:
        return []
    return 줄목록


# ────────────────────────────────────────
# ② 에러 로그만 필터링
# ────────────────────────────────────────
def 에러_필터링(로그목록):
    """
    전체 로그에서 ERROR 가 포함된 줄만 추출
    """
    에러목록 = []
    for 로그 in 로그목록:
        로그 = 로그.strip()
        if "ERROR" in 로그:
            에러목록.append(로그)
    return 에러목록


# ────────────────────────────────────────
# ③ 버그 ID / 시간 추출
# ────────────────────────────────────────
def 버그정보_추출(로그):
    """
    로그 한 줄에서 버그 ID와 발생 시간을 정규표현식으로 추출
    """
    버그ID = re.search(r"BUG-\d+", 로그)
    시간   = re.search(r"\d{2}:\d{2}:\d{2}", 로그)
    return {
        "로그"  : 로그,
        "버그ID": 버그ID.group() if 버그ID else "없음",
        "시간"  : 시간.group()   if 시간   else "없음"
    }


# ────────────────────────────────────────
# ④ 심각도 자동 분류
# ────────────────────────────────────────
def 심각도_분류(로그):
    """
    로그 내용의 키워드로 심각도를 자동 판단
    Critical / Major / Minor 중 하나 반환

    Major는 공통 색상에 없으므로 내부 매핑 유지
    """
    키워드맵 = {
        "Critical": ["충돌", "crash", "서버 다운", "응답 없음"],
        "Major"   : ["프레임 드랍", "렉", "지연", "오류"],
        "Minor"   : ["경고", "warning", "메모리"],
    }
    for 심각도, 키워드목록 in 키워드맵.items():
        for 키워드 in 키워드목록:
            if 키워드.lower() in 로그.lower():
                return 심각도
    return "Minor"


# ────────────────────────────────────────
# ⑤ 중복 버그 감지
# ────────────────────────────────────────
def 중복_감지(버그목록):
    """
    같은 버그 ID가 2회 이상 발생한 항목 탐지
    """
    ID목록  = [버그["버그ID"] for 버그 in 버그목록 if 버그["버그ID"] != "없음"]
    카운터  = Counter(ID목록)
    중복목록 = {id: 수 for id, 수 in 카운터.items() if 수 > 1}
    return 중복목록


# ────────────────────────────────────────
# ⑥ 이전 결과와 비교
# ────────────────────────────────────────
def 히스토리_비교(현재버그목록):
    """
    이전 실행 결과와 비교해서 신규/해결 버그 추적
    bug_history.json 사용
    """
    히스토리파일 = 기준경로 / "결과" / "bug_history.json"
    현재ID목록   = set(버그["버그ID"] for 버그 in 현재버그목록)

    if not 히스토리파일.exists():
        히스토리_저장(현재버그목록)
        return {"신규": list(현재ID목록), "해결": []}

    with open(히스토리파일, "r", encoding="utf-8") as f:
        이전데이터 = json.load(f)

    이전ID목록 = set(이전데이터.get("버그ID목록", []))

    신규버그 = list(현재ID목록 - 이전ID목록)
    해결버그 = list(이전ID목록 - 현재ID목록)

    히스토리_저장(현재버그목록)
    return {"신규": 신규버그, "해결": 해결버그}


def 히스토리_저장(버그목록):
    """
    현재 버그 목록을 JSON 파일로 저장
    저장 위치: 결과/bug_history.json
    """
    결과폴더 = 결과폴더_생성(기준경로)

    히스토리파일 = 결과폴더 / "bug_history.json"
    오늘        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    데이터 = {
        "날짜"     : 오늘,
        "버그ID목록": [버그["버그ID"] for 버그 in 버그목록]
    }

    with open(히스토리파일, "w", encoding="utf-8") as f:
        json.dump(데이터, f, ensure_ascii=False, indent=4)


# ────────────────────────────────────────
# ⑦ 리포트 저장
# ────────────────────────────────────────
def 엑셀_저장(버그목록, 전체로그, 비교결과, 중복결과):
    """
    분석 결과를 5시트 엑셀 파일로 저장
    공통 모듈 사용: 색상, 헤더_스타일, 행_색상, 열너비_조정

    저장 위치: 결과/bug_report_YYYY-MM-DD.xlsx

    시트 구성:
    - 전체 로그   : ERROR🔴 / WARNING🟡 / INFO🟢 색상 구분
    - 버그 리포트 : Critical🔴 / Major🟡 / Minor🟢 심각도 분류
    - 통계 요약   : 유형별 건수
    - 변경 사항   : 신규🔴 / 해결🟢 버그
    - 중복 버그   : 반복 발생 버그🟡
    """
    결과폴더 = 결과폴더_생성(기준경로)

    시각   = 타임스탬프()
    파일명 = 결과폴더 / f"bug_report_{시각}.xlsx"

    wb = Workbook()

    # ── 시트 1: 전체 로그 ──
    ws1 = wb.active
    ws1.title = "전체 로그"
    ws1.append(["로그 유형", "내용"])
    헤더_스타일(ws1)

    for 로그 in 전체로그:
        로그 = 로그.strip()
        유형 = "INFO"
        if "ERROR"   in 로그: 유형 = "ERROR"
        elif "WARNING" in 로그: 유형 = "WARNING"

        ws1.append([유형, 로그])
        현재행 = ws1.max_row
        배경색 = 색상_가져오기(유형)
        for 셀 in ws1[현재행]:
            셀.fill = PatternFill("solid", fgColor=배경색)

    # ── 시트 2: 버그 리포트 ──
    ws2 = wb.create_sheet(title="버그 리포트")
    ws2.append(["버그ID", "시간", "심각도", "로그 내용"])
    헤더_스타일(ws2)

    # 심각도별 색상 (Major는 공통 색상에 없으므로 별도 매핑)
    심각도색상 = {
        "Critical": 색상_가져오기("Critical"),
        "Major"   : 색상_가져오기("Medium"),
        "Minor"   : 색상_가져오기("Low"),
    }

    for 버그 in 버그목록:
        심각도 = 심각도_분류(버그["로그"])
        ws2.append([버그["버그ID"], 버그["시간"], 심각도, 버그["로그"]])
        현재행 = ws2.max_row
        배경색 = 심각도색상[심각도]
        for 셀 in ws2[현재행]:
            셀.fill = PatternFill("solid", fgColor=배경색)
            셀.font = Font(color="FFFFFF")

    # ── 시트 3: 통계 요약 ──
    ws3 = wb.create_sheet(title="통계 요약")
    ws3.append(["항목", "건수"])
    헤더_스타일(ws3)

    통계데이터 = [
        ("ERROR",   sum(1 for 로그 in 전체로그 if "ERROR"   in 로그)),
        ("WARNING", sum(1 for 로그 in 전체로그 if "WARNING" in 로그)),
        ("INFO",    sum(1 for 로그 in 전체로그 if "INFO"    in 로그)),
        ("전체",    len(전체로그)),
    ]

    for 항목, 건수 in 통계데이터:
        ws3.append([항목, 건수])
        현재행 = ws3.max_row
        배경색 = 색상_가져오기(항목) if 항목 != "전체" else 색상_가져오기("헤더")
        행_색상(ws3, 현재행, 배경색)

    # ── 시트 4: 변경 사항 ──
    ws4 = wb.create_sheet(title="변경 사항")
    ws4.append(["구분", "버그ID"])
    헤더_스타일(ws4)

    for id in 비교결과["신규"]:
        ws4.append(["🆕 신규", id])
        현재행 = ws4.max_row
        for 셀 in ws4[현재행]:
            셀.fill = PatternFill("solid", fgColor=색상_가져오기("Critical"))
            셀.font = Font(color="FFFFFF")

    for id in 비교결과["해결"]:
        ws4.append(["✅ 해결", id])
        현재행 = ws4.max_row
        for 셀 in ws4[현재행]:
            셀.fill = PatternFill("solid", fgColor=색상_가져오기("해결"))
            셀.font = Font(color="FFFFFF")

    # ── 시트 5: 중복 버그 ──
    ws5 = wb.create_sheet(title="중복 버그")
    ws5.append(["버그ID", "발생 횟수"])
    헤더_스타일(ws5)

    for id, 횟수 in 중복결과.items():
        ws5.append([id, 횟수])
        현재행 = ws5.max_row
        행_색상(ws5, 현재행, "FFC000")  # 노랑

    # 모든 시트 열 너비 자동 조정
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        열너비_조정(ws)

    wb.save(파일명)
    Latest_복사(파일명, "bug_report")
    return 파일명

def 리포트_JSON_저장(버그목록, 전체로그, 비교결과, 중복결과):
    """
    분석 결과를 JSON 파일로 저장 (자동화 연동용)
    """
    결과폴더 = 결과폴더_생성(기준경로)
    시각   = 타임스탬프()
    파일명 = 결과폴더 / f"bug_report_{시각}.json"

    데이터 = {
        "분석_일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "통계": {
            "총_로그": len(전체로그),
            "ERROR": sum(1 for 로그 in 전체로그 if "ERROR" in 로그),
            "WARNING": sum(1 for 로그 in 전체로그 if "WARNING" in 로그),
            "INFO": sum(1 for 로그 in 전체로그 if "INFO" in 로그),
        },
        "버그_목록": [
            {
                "버그ID": 버그["버그ID"],
                "시간": 버그["시간"],
                "심각도": 심각도_분류(버그["로그"]),
                "로그": 버그["로그"]
            }
            for 버그 in 버그목록
        ],
        "변경_사항": {
            "신규": 비교결과["신규"],
            "해결": 비교결과["해결"]
        },
        "중복_버그": 중복결과
    }

    JSON_쓰기(파일명, 데이터)
    Latest_복사(파일명, "bug_report")
    return 파일명


def 리포트_HTML_저장(버그목록, 전체로그, 비교결과, 중복결과):
    """
    분석 결과를 HTML 리포트로 저장 (웹 브라우저용)
    """
    결과폴더 = 결과폴더_생성(기준경로)
    시각   = 타임스탬프()
    파일명 = 결과폴더 / f"bug_report_{시각}.html"

    error수   = sum(1 for 로그 in 전체로그 if "ERROR"   in 로그)
    warning수 = sum(1 for 로그 in 전체로그 if "WARNING" in 로그)
    info수    = sum(1 for 로그 in 전체로그 if "INFO"    in 로그)

    # 요약
    요약 = f"""
    <div class="summary">
        <h2>📊 분석 결과</h2>
        <p><strong>분석 일시:</strong> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        <p><strong>총 로그 수:</strong> {len(전체로그)}건</p>
        <p><strong>ERROR:</strong> <span class="미해결">{error수}건</span></p>
        <p><strong>WARNING:</strong> <span class="진행중">{warning수}건</span></p>
        <p><strong>INFO:</strong> <span class="해결">{info수}건</span></p>
    </div>
    """

    # 버그 리포트
    버그리포트 = "<h2>🐛 버그 리포트</h2>"
    if 버그목록:
        버그리포트 += "<table><tr><th>버그ID</th><th>시간</th><th>심각도</th><th>로그 내용</th></tr>"
        for 버그 in 버그목록:
            심각도 = 심각도_분류(버그["로그"])
            클래스 = {"Critical": "Critical", "Major": "High", "Minor": "Low"}.get(심각도, "")
            버그리포트 += f'<tr><td>{버그["버그ID"]}</td><td>{버그["시간"]}</td><td class="{클래스}">{심각도}</td><td>{버그["로그"]}</td></tr>'
        버그리포트 += "</table>"
    else:
        버그리포트 += "<p>에러 로그가 없습니다 🎉</p>"

    # 변경 사항
    변경 = "<h2>🔄 변경 사항</h2>"
    if 비교결과["신규"] or 비교결과["해결"]:
        변경 += "<table><tr><th>구분</th><th>버그ID</th></tr>"
        for id in 비교결과["신규"]:
            변경 += f'<tr><td class="미해결">🆕 신규</td><td>{id}</td></tr>'
        for id in 비교결과["해결"]:
            변경 += f'<tr><td class="해결">✅ 해결</td><td>{id}</td></tr>'
        변경 += "</table>"
    else:
        변경 += "<p>변경 사항이 없습니다.</p>"

    # 중복 버그
    중복 = "<h2>⚠️ 중복 버그</h2>"
    if 중복결과:
        중복 += "<table><tr><th>버그ID</th><th>발생 횟수</th></tr>"
        for id, 횟수 in 중복결과.items():
            중복 += f'<tr><td>{id}</td><td class="진행중">{횟수}회</td></tr>'
        중복 += "</table>"
    else:
        중복 += "<p>중복 버그가 없습니다.</p>"

    본문 = 요약 + 버그리포트 + 변경 + 중복

    HTML_쓰기(파일명, "QA 로그 분석 리포트", 본문)
    Latest_복사(파일명, "bug_report")
    return 파일명