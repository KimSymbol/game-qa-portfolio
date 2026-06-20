# 역할: JSON 사양 파일을 기반으로 ISTQB 테스트 설계 기법을 자동화
#       → 동등분할 / 경계값 분석 / 결정 테이블 기반 TC 자동 생성
#
# ISTQB 기법:
#   - 동등분할 (Equivalence Partitioning)
#   - 경계값 분석 (Boundary Value Analysis)
#   - 결정 테이블 (Decision Table)
#
# 사용법:
#   from tc_generator import 전체_TC_생성
#   결과 = 전체_TC_생성("specs/login.json")

import sys
import json
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from common.excel_style import 헤더_스타일, 행_색상, 색상_가져오기, 열너비_조정

# 02.tools 폴더를 Python 경로에 추가
sys.path.insert(0, str(Path(__file__).parent.parent))

from common.file_io import 결과폴더_생성, 타임스탬프, Latest_복사
from common.logger import 로거_생성

log = 로거_생성("tc-generator")

# tc_generator.py 가 있는 폴더를 기준 경로로 설정
기준경로 = Path(__file__).parent


# ────────────────────────────────────────
# 깡통 템플릿 생성
# ────────────────────────────────────────
def 템플릿_생성(형식="csv"):
    """
    헤더만 있는 빈 TC 파일 생성
    QA가 직접 수기로 작성할 때 사용

    매개변수:
    - 형식: "csv" 또는 "xlsx"

    저장 위치: 결과/tc_template.csv / .xlsx
    """
    결과폴더 = 결과폴더_생성(기준경로)

    헤더 = [
        "TC_ID", "테스트명", "분류", "전제조건",
        "테스트단계", "예상결과", "실제결과", "결과",
        "심각도", "우선순위", "플랫폼", "발견자", "발견일"
    ]

    if 형식 == "csv":
        파일명 = 결과폴더 / "tc_template.csv"
        with open(파일명, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(헤더)
        log.info(f"TC 템플릿 생성: {파일명}")

    elif 형식 == "xlsx":
        파일명 = 결과폴더 / "tc_template.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "테스트 케이스"
        ws.append(헤더)
        헤더_스타일(ws)
        열너비_조정(ws, 최대너비=50)
        wb.save(파일명)
        log.info(f"TC 템플릿 생성: {파일명}")

    else:
        log.error(f"지원하지 않는 형식: {형식}")
        return None

    return 파일명


# ────────────────────────────────────────
# ① 사양 파일 로딩
# ────────────────────────────────────────
def 사양_로딩(파일명):
    """
    JSON 사양 파일을 로딩

    매개변수:
    - 파일명: specs/ 폴더의 JSON 파일 경로

    반환값:
    - 사양 딕셔너리
    - None: 파일이 없거나 형식 오류
    """
    경로 = Path(파일명)
    if not 경로.exists():
        경로 = 기준경로 / "specs" / 파일명

    if not 경로.exists():
        log.error(f"사양 파일 없음: {파일명}")
        return None

    try:
        with open(경로, "r", encoding="utf-8") as f:
            사양 = json.load(f)
        log.info(f"사양 로딩 완료: {경로.name} ({사양.get('기능명', '')})")
        return 사양
    except json.JSONDecodeError as e:
        log.error(f"JSON 형식 오류: {경로} - {e}")
        return None
    except Exception as e:
        log.error(f"사양 로딩 실패: {type(e).__name__}: {e}")
        return None


# ────────────────────────────────────────
# ② 동등분할 TC 생성
# ────────────────────────────────────────
def 동등분할_TC_생성(사양):
    """
    동등 분할 기법으로 TC 생성
    유효/무효 입력값마다 TC 1개씩 생성
    """
    TC목록 = []
    동등분할 = 사양.get("동등분할", {})

    for 파라미터명, 값분류 in 동등분할.items():
        # 유효한 값들
        for 값 in 값분류.get("유효", []):
            TC목록.append({
                "테스트명": f"{파라미터명} 정상값 '{값}' 입력 시 동작 확인",   # ← 접두사 제거
                "파라미터": 파라미터명,
                "입력값": 값,
                "구분": "유효",
                "예상결과": f"{파라미터명}이(가) 정상 처리됨",
                "심각도": "Medium",
                "우선순위": "Medium"
            })
        # 무효한 값들
        for 값 in 값분류.get("무효", []):
            TC목록.append({
                "테스트명": f"{파라미터명}에 비정상값 '{값}' 입력 시 오류 처리 확인",   # ← 접두사 제거
                "파라미터": 파라미터명,
                "입력값": 값,
                "구분": "무효",
                "예상결과": f"{파라미터명} 오류 메시지 표시",
                "심각도": "High",
                "우선순위": "Medium"
            })

    return TC목록


# ────────────────────────────────────────
# ③ 경계값 TC 생성
# ────────────────────────────────────────
def 경계값_TC_생성(사양):
    """경계값 분석 기법으로 TC 생성"""
    TC목록 = []
    경계값 = 사양.get("경계값", {})

    for 파라미터명, 범위 in 경계값.items():
        최소 = 범위.get("최소")
        최대 = 범위.get("최대")
        단위 = 범위.get("단위", "")

        if 최소 is None or 최대 is None:
            continue

        # 접두사 제거된 테스트명
        경계점목록 = [
            (최소 - 1, "최소 미만",   f"{파라미터명} 최소 미만값({최소 - 1}{단위}) 입력 시 오류 처리 확인",   "오류 메시지 표시", "High"),
            (최소,      "최소값",      f"{파라미터명} 최소값({최소}{단위}) 입력 시 정상 동작 확인",          "정상 처리",        "Medium"),
            (최대,      "최대값",      f"{파라미터명} 최대값({최대}{단위}) 입력 시 정상 동작 확인",          "정상 처리",        "Medium"),
            (최대 + 1, "최대 초과",   f"{파라미터명} 최대 초과값({최대 + 1}{단위}) 입력 시 오류 처리 확인", "오류 메시지 표시", "High"),
        ]

        for 값, 구분, 테스트명, 예상결과, 심각도 in 경계점목록:
            TC목록.append({
                "테스트명": 테스트명,
                "파라미터": 파라미터명,
                "입력값": f"{값}{단위}",
                "구분": 구분,
                "예상결과": 예상결과,
                "심각도": 심각도,
                "우선순위": "High"
            })

    return TC목록


# ────────────────────────────────────────
# ④ 결정테이블 TC 생성
# ────────────────────────────────────────
def 결정테이블_TC_생성(사양):
    """결정 테이블 기법으로 TC 생성"""
    TC목록 = []
    결정테이블 = 사양.get("결정테이블", [])

    for 항목 in 결정테이블:
        조건 = 항목.get("조건", {})

        # 조건을 자연어 형태로 표현
        # 예: "ID=유효, PW=유효" → "정상 ID + 정상 PW"
        조건요약 = []
        for k, v in 조건.items():
            상태 = "정상" if v == "유효" else "비정상"
            조건요약.append(f"{상태} {k}")
        조건문자열 = " + ".join(조건요약)

        TC목록.append({
            "테스트명": f"{조건문자열} 조합 시 동작 확인",   # ← 접두사 제거 + 자연어화
            "파라미터": " + ".join(조건.keys()),
            "입력값": ", ".join([f"{k}={v}" for k, v in 조건.items()]),
            "구분": "조합",
            "예상결과": 항목.get("예상결과", ""),
            "심각도": 항목.get("심각도", "Medium"),
            "우선순위": 항목.get("우선순위", "Medium")
        })

    return TC목록


# ────────────────────────────────────────
# ⑤ 재현 절차 포맷팅
# ────────────────────────────────────────
def 재현절차_포맷(공통헤더, 테스트액션):
    """
    공통 헤더 + 테스트별 액션을 마크다운 번호 목록으로 변환

    매개변수:
    - 공통헤더 : 공통 재현 절차 (리스트)
    - 테스트액션: 이 TC의 마지막 액션 (문자열)

    반환값:
    - 번호가 매겨진 절차 문자열
    """
    전체 = list(공통헤더) + [테스트액션]
    return "\n".join([f"{i+1}. {단계}" for i, 단계 in enumerate(전체)])


# ────────────────────────────────────────
# ⑥ TC → testcases.csv 변환
# ────────────────────────────────────────
def CSV_저장(사양, TC목록):
    """
    TC 목록을 testcases.csv 형식으로 저장
    기법_컬럼_포함 설정에 따라 컬럼 구성 변경
    """
    결과폴더 = 결과폴더_생성(기준경로)

    기능명 = 사양.get("기능명", "unknown")
    분류   = 사양.get("분류", "")
    전제조건 = 사양.get("전제조건", "")
    공통헤더 = 사양.get("공통_재현절차_헤더", [])

    시각 = 타임스탬프()
    파일명 = 결과폴더 / f"testcases_{기능명}_{시각}.csv"

    # 기법 컬럼 포함 여부에 따라 헤더 구성
    헤더 = [
        "TC_ID", "테스트명", "분류", "전제조건",
        "테스트단계", "예상결과", "실제결과", "결과",
        "심각도", "우선순위", "플랫폼", "발견자", "발견일"
    ]
    if 기법_컬럼_포함:
        헤더.append("기법")

    with open(파일명, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(헤더)

        for i, TC in enumerate(TC목록, 1):
            테스트액션 = f"{TC['파라미터']} 에 '{TC['입력값']}' 입력 후 동작 수행"
            재현절차 = 재현절차_포맷(공통헤더, 테스트액션)

            행 = [
                f"TC-{i:03d}",
                TC["테스트명"],
                분류,
                전제조건,
                재현절차,
                TC["예상결과"],
                "",
                "",
                TC["심각도"],
                TC["우선순위"],
                "PC",
                "자동생성",
                datetime.now().strftime("%Y-%m-%d"),
            ]
            if 기법_컬럼_포함:
                행.append(TC["기법"])

            writer.writerow(행)

    Latest_복사(파일명, f"testcases_{기능명}")
    log.info(f"CSV 저장 완료: {파일명} ({len(TC목록)}건)")
    return 파일명

# ────────────────────────────────────────
# ⑥-2. TC → testcases.xlsx 변환
# ────────────────────────────────────────
def XLSX_저장(사양, TC목록):
    """TC 목록을 xlsx 형식으로 저장"""
    결과폴더 = 결과폴더_생성(기준경로)

    기능명 = 사양.get("기능명", "unknown")
    분류   = 사양.get("분류", "")
    전제조건 = 사양.get("전제조건", "")
    공통헤더 = 사양.get("공통_재현절차_헤더", [])

    시각   = 타임스탬프()
    파일명 = 결과폴더 / f"testcases_{기능명}_{시각}.xlsx"

    wb = Workbook()

    # ── 시트 1: 전체 TC 목록 ──
    ws1 = wb.active
    ws1.title = "전체 TC"

    헤더 = [
        "TC_ID", "테스트명", "분류", "전제조건",
        "테스트단계", "예상결과", "실제결과", "결과",
        "심각도", "우선순위", "플랫폼", "발견자", "발견일"
    ]
    ws1.append(헤더)
    헤더_스타일(ws1)

    for i, TC in enumerate(TC목록, 1):
        테스트액션 = f"{TC['파라미터']} 에 '{TC['입력값']}' 입력 후 동작 수행"
        재현절차 = 재현절차_포맷(공통헤더, 테스트액션)

        ws1.append([
            f"TC-{i:03d}",
            TC["테스트명"],
            분류,
            전제조건,
            재현절차,
            TC["예상결과"],
            "",
            "",
            TC["심각도"],
            TC["우선순위"],
            "PC",
            "자동생성",
            datetime.now().strftime("%Y-%m-%d"),
        ])

    # ── 시트 2: 사양 요약 ──
    ws2 = wb.create_sheet(title="사양 요약")
    ws2.append(["항목", "내용"])
    헤더_스타일(ws2)

    ws2.append(["기능명", 기능명])
    ws2.append(["분류", 분류])
    ws2.append(["전제조건", 전제조건])
    ws2.append(["생성 시각", 시각])
    ws2.append(["총 TC 수", str(len(TC목록))])

    for ws in [ws1, ws2]:
        열너비_조정(ws, 최대너비=50)

    wb.save(파일명)
    Latest_복사(파일명, f"testcases_{기능명}")
    log.info(f"XLSX 저장 완료: {파일명} ({len(TC목록)}건)")
    return 파일명


# ────────────────────────────────────────
# ⑦ 전체 TC 생성 (메인)
# ────────────────────────────────────────
def 전체_TC_생성(사양파일):
    """사양 파일로부터 전체 TC를 자동 생성"""
    사양 = 사양_로딩(사양파일)
    if 사양 is None:
        return None

    기능명 = 사양.get("기능명", "unknown")
    log.info(f"TC 생성 시작: {기능명}")

    동등분할_TC = 동등분할_TC_생성(사양)
    경계값_TC   = 경계값_TC_생성(사양)
    결정테이블_TC = 결정테이블_TC_생성(사양)

    전체 = 동등분할_TC + 경계값_TC + 결정테이블_TC

    csv경로  = CSV_저장(사양, 전체)
    xlsx경로 = XLSX_저장(사양, 전체)

    결과 = {
        "csv"  : csv경로,
        "xlsx" : xlsx경로,
        "건수" : len(전체),
    }

    log.info(f"TC 생성 완료: {기능명} ({len(전체)}건)")
    return 결과