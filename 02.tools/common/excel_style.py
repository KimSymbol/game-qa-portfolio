# 역할: 모든 도구에서 공통으로 사용하는 엑셀 스타일 함수
#
# 사용법:
#   from common.excel_style import 헤더_스타일, 열너비_조정, 색상, 색상_가져오기

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ── 공통 색상 정의 ──
# 새 값을 추가하면 모든 도구에 자동 반영
색상 = {
    # 심각도
    "Critical": "FF0000",
    "High"    : "FF6600",
    "Medium"  : "FFC000",
    "Low"     : "70AD47",
    # 상태
    "해결"    : "70AD47",
    "진행중"  : "FFC000",
    "미해결"  : "FF0000",
    # 테스트 결과
    "Pass"    : "70AD47",
    "Fail"    : "FF0000",
    "Block"   : "FFC000",
    "Skip"    : "BFBFBF",
    # 로그 유형
    "ERROR"   : "FF0000",
    "WARNING" : "FFC000",
    "INFO"    : "70AD47",
    # 우선순위
    "High_P"  : "FF0000",
    "Medium_P": "FFC000",
    "Low_P"   : "70AD47",
    # 공통
    "헤더"    : "4472C4",
    "기본"    : "BFBFBF",
}

# 우선순위 전용 색상 (키 이름 충돌 방지용)
우선순위_색상 = {
    "High"  : "FF0000",
    "Medium": "FFC000",
    "Low"   : "70AD47",
}


def 헤더_스타일(ws, 행번호=1):
    """
    지정한 행을 헤더 스타일로 꾸밈
    배경: 파란색 / 글씨: 흰색 굵게 / 정렬: 가운데
    """
    for 셀 in ws[행번호]:
        셀.fill      = PatternFill("solid", fgColor=색상["헤더"])
        셀.font      = Font(color="FFFFFF", bold=True)
        셀.alignment = Alignment(horizontal="center")


def 셀_색상(셀, 배경색, 글자색="FFFFFF", 굵게=True, 가운데=True):
    """
    개별 셀에 스타일 적용
    """
    셀.fill = PatternFill("solid", fgColor=배경색)
    셀.font = Font(color=글자색, bold=굵게)
    if 가운데:
        셀.alignment = Alignment(horizontal="center")


def 행_색상(ws, 행번호, 배경색, 글자색="FFFFFF", 굵게=True):
    """
    행 전체에 동일한 스타일 적용
    """
    for 셀 in ws[행번호]:
        셀.fill      = PatternFill("solid", fgColor=배경색)
        셀.font      = Font(color=글자색, bold=굵게)
        셀.alignment = Alignment(horizontal="center")


def 색상_가져오기(값, 기본색="BFBFBF"):
    """
    값에 해당하는 색상 코드 반환
    없는 값이면 기본 회색 반환
    """
    return 색상.get(값, 기본색)


def 열너비_조정(ws, 최대너비=60):
    """
    워크시트의 각 열 너비를 내용 길이에 맞게 자동 조정
    병합 셀(MergedCell) 에러 방지 포함
    """
    for i, 열 in enumerate(ws.columns, 1):
        최대길이 = 0
        열이름  = get_column_letter(i)
        for 셀 in 열:
            try:
                if 셀.value:
                    최대길이 = max(최대길이, len(str(셀.value)))
            except:
                pass
        ws.column_dimensions[열이름].width = min(최대길이 + 4, 최대너비)